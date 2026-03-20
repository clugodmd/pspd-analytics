#!/usr/bin/env python3
"""
payroll_refresh.py — PSPD Doctor Payroll Data Pipeline
======================================================
Generates data/payroll.json consumed by payroll.html.

THREE-STATE PERIOD LIFECYCLE
─────────────────────────────
  live        Pay period is active (today ≤ period_end). Data recalculates
              on every Azure run (twice daily). Doctors actively seeing patients.

  pending     Period ended; snapshot was taken Sunday at 5am CST (after the
              Saturday morning Denticon dump has settled). State is COMPLETELY
              FROZEN — zero further updates. Tanya reviews before paying.
              payroll.json "snapshot" is the authoritative frozen copy.
              NO "live" drift comparison — pending is immutable.

  locked      Tanya/Dr. Lugo reviewed and clicked "Lock Period" in the dashboard
              (PIN-confirmed). Means payroll was actually PAID. Data is
              permanently frozen. Script will NEVER overwrite a locked period.

TRANSITION TRIGGERS
────────────────────
  live → pending    : Script runs Sunday at 5am CST. Detects period_end < today.
                      Auto-snapshots current doctor/office data into
                      period.snapshot, sets state = "pending".
                      NEVER updates a pending period again after this snapshot.

  pending → locked  : Front-end "Lock Period" button (PIN required).
                      Means payroll was actually PAID.
                      Sets state = "locked", lockedAt, lockedBy in JSON.
                      Script respects this — locked periods are read-only.

PAY CALCULATION RULES
──────────────────────
  All associates:  pay = (collections - xray_collections) × rate
  Dr. Schrack:     pay = collections × rate  (1099 oral surgeon, no x-rays)
  Dr. Lugo (owner): pay = collections × rate  (owner flag, salary basis)

  X-ray exclusion applies to: procedures starting with D02* and D03*
  Dr. Schrack is exempt (xray_exempt = True in config): he does not
  perform diagnostic x-rays — his procedures are surgical. His pay_basis
  is 'collections_no_xray_exempt', which means payNo = payWith always.

  ROUNDING: Round only at the final output step. Accumulate intermediate
  sums as full-precision floats. This eliminates the ~0.02% gap vs Tanya's
  spreadsheet caused by rounding at each sub-step.

DATA VIEWS
───────────
  rpt.vw_income_allocation           ← THIS is Tanya's "Income Allocation
                                       Report - Detail" from Denticon. This
                                       is the source of truth for all pay
                                       calculations. Use this view.

  rpt.vw_doctor_collections_no_xray  ← Pre-computed x-ray exclusion view.
                                       NOT used — we compute x-ray exclusion
                                       ourselves from transaction-level data
                                       so we can audit individual procedures.
                                       Using this pre-computed view was likely
                                       the source of the 0.02% accuracy gap.

  rpt.vw_pay_periods                 ← Pay period definitions (biweekly).

Usage:
  # Automated (Azure Task, runs 1am Saturday):
  python scripts/payroll_refresh.py

  # Force process specific period:
  python scripts/payroll_refresh.py --period 3.27.26

  # All historical periods (re-process non-locked):
  python scripts/payroll_refresh.py --all-periods

  # From Tanya's Excel export (manual override):
  python scripts/payroll_refresh.py --from-excel path/to/Denticon_NewMonthlyIncAllD.xlsx

  # List available pay periods:
  python scripts/payroll_refresh.py --list-periods

Environment Variables:
  AZURE_SQL_CONN_STR — Full ODBC connection string for Denticon Azure SQL
"""

import json
import os
import sys
import argparse
import re
from datetime import datetime, date, timedelta
from pathlib import Path


# ─────────────────────────────────────────────────────────────────────────────
# DOCTOR CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
#
# pay_basis options:
#   'collections'             — W-2 associate: pay = (coll - xray) × rate
#   'collections_xray_exempt' — 1099 oral surgeon (Schrack): pay = coll × rate
#                               He performs surgery, not diagnostic x-rays.
#                               x-ray exclusion does not apply to his work.
#   'salary'                  — Owner (Dr. Lugo): pay = coll × rate
#                               Owner always uses payWith formula (no exclusion)
#
# The 'xray_exempt' flag is the authoritative check in process_transactions().
# It can be set either via pay_basis == 'collections_xray_exempt' or
# explicitly via owner == True. Both bypass x-ray exclusion.

DOCTOR_CONFIG = {
    # Denticon "Last, First" → pay config
    # Must match provider name in PGID4951_PROVIDER exactly
    'Slaven, Chad':       {'display': 'Dr. Slaven',  'pct': 0.36, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Menon, Leena':       {'display': 'Dr. Menon',   'pct': 0.35, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Choong, Carissa':    {'display': 'Dr. Choong',  'pct': 0.35, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Benton, Patricia':   {'display': 'Dr. Benton',  'pct': 0.33, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Welter, Erin':       {'display': 'Dr. Welter',  'pct': 0.31, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Patel, Dusayant':    {'display': 'Dr. Patel',   'pct': 0.32, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    'Bell, Kendra':       {'display': 'Dr. Bell',    'pct': 0.35, 'owner': False, 'pay_basis': 'collections',             'xray_exempt': False},
    # Dr. Schrack: 1099 oral surgeon — wisdom teeth & limited oral surgery only.
    # He does NOT perform diagnostic x-rays. His procedures are surgical (D7xxx).
    # X-ray exclusion does NOT apply → payNo = payWith always.
    'Schrack, Donald':    {'display': 'Dr. Schrack', 'pct': 0.45, 'owner': False, 'pay_basis': 'collections_xray_exempt', 'xray_exempt': True},
    # Dr. Lugo: Owner. Always uses collections × rate (payWith formula).
    # payNo is set to 0.0 by convention (x-ray savings shown elsewhere).
    'Lugo, Christopher':  {'display': 'Dr. Lugo',    'pct': 0.36, 'owner': True,  'pay_basis': 'salary',                  'xray_exempt': True},
}

# Terminated / inactive doctors — tracked for residual collections display
TERMED_DOCTORS = {
    'Kirk, Kyle':  {'display': 'Dr. Kirk',  'note': 'Terminated'},
    'Ping, Sita':  {'display': 'Dr. Ping',  'note': 'Terminated'},
    'Laws':        {'display': 'Dr. Laws',  'note': 'Terminated'},
}

# Provider ID → "Last, First" (auto-populated from PGID4951_PROVIDER)
PROVIDER_ID_MAP = {}

# Office ID → display name (auto-populated from PGID4951_OFFICE)
OFFICE_ID_MAP = {}

# Office name text matching (for Excel path and raw SQL strings)
OFFICE_NAME_MAP = {
    'EVERETT':      ('Everett',      'EV'),
    'LAKE STEVENS': ('Lake Stevens', 'LS'),
    'MARYSVILLE':   ('Marysville',  'MV'),
    'MONROE':       ('Monroe',      'MO'),
    'STANWOOD':     ('Stanwood',    'SW'),
}

# X-ray procedure code prefixes — excluded from non-exempt doctor pay
# D02* = diagnostic radiographic imaging (bitewings, periapicals, panoramic)
# D03* = diagnostic imaging (cone beam CT)
XRAY_PREFIXES = ('D02', 'D03')

# Hardcoded fallback pay period schedule (for Excel mode when DB unavailable)
FALLBACK_PAY_PERIODS = [
    ('1.16.26',  '2025-12-27', '2026-01-09', '2026-01-16'),
    ('1.30.26',  '2026-01-10', '2026-01-23', '2026-01-30'),
    ('2.13.26',  '2026-01-24', '2026-02-06', '2026-02-13'),
    ('2.27.26',  '2026-02-07', '2026-02-20', '2026-02-27'),
    ('3.13.26',  '2026-02-21', '2026-03-06', '2026-03-13'),
    ('3.27.26',  '2026-03-07', '2026-03-20', '2026-03-27'),
    ('4.10.26',  '2026-03-21', '2026-04-03', '2026-04-10'),
    ('4.24.26',  '2026-04-04', '2026-04-17', '2026-04-24'),
    ('5.8.26',   '2026-04-18', '2026-05-01', '2026-05-08'),
    ('5.22.26',  '2026-05-02', '2026-05-15', '2026-05-22'),
]


# ─────────────────────────────────────────────────────────────────────────────
# LOAD DOCTOR CONFIG FROM doctors.json
# ─────────────────────────────────────────────────────────────────────────────

def load_doctor_config():
    """
    Load doctor config from data/doctors.json.
    Falls back to hardcoded DOCTOR_CONFIG if file missing.
    Returns (active_dict, termed_dict).
    """
    global DOCTOR_CONFIG, TERMED_DOCTORS

    candidates = [
        Path(__file__).resolve().parent.parent / "data" / "doctors.json",
        Path(__file__).resolve().parent / "data" / "doctors.json",
        Path("data/doctors.json"),
    ]

    doctors_json = None
    for p in candidates:
        if p.exists():
            doctors_json = p
            break

    if doctors_json is None:
        print("INFO: data/doctors.json not found — using hardcoded DOCTOR_CONFIG")
        return DOCTOR_CONFIG, TERMED_DOCTORS

    try:
        with open(doctors_json, 'r') as f:
            data = json.load(f)

        active = {}
        for name, doc in data.get("doctors", {}).items():
            pay_basis = doc.get('pay_basis', 'collections')
            owner = doc.get('owner', False)
            # xray_exempt: explicit flag wins; fallback from pay_basis/owner
            xray_exempt = doc.get('xray_exempt',
                pay_basis == 'collections_xray_exempt' or owner)
            active[name] = {
                'display':    doc['display'],
                'pct':        doc['pct'],
                'owner':      owner,
                'pay_basis':  pay_basis,
                'xray_exempt': xray_exempt,
            }

        termed = {}
        for name, doc in data.get("terminated", {}).items():
            termed[name] = {
                'display': doc['display'],
                'note':    doc.get('note', 'Terminated'),
            }

        DOCTOR_CONFIG = active
        TERMED_DOCTORS = termed
        print(f"  Loaded {len(active)} active + {len(termed)} terminated doctors from {doctors_json.name}")
        return active, termed

    except Exception as e:
        print(f"WARNING: Failed to load doctors.json: {e} — using hardcoded config")
        return DOCTOR_CONFIG, TERMED_DOCTORS


# ─────────────────────────────────────────────────────────────────────────────
# HARDCODED HISTORICAL DATA — Tanya's verified paid amounts (locked forever)
# ─────────────────────────────────────────────────────────────────────────────
# Source: 'NEW_MASTER PAY TEMPLATE ASSOCIATE' Excel file, 'No X-Rays' tabs.
# These are the EXACT amounts that were paid. They are the ground truth.
# The pipeline will NEVER re-query or overwrite these. They are injected
# into payroll.json with state = "locked", hardcoded = True.

def get_hardcoded_periods():
    """Return Tanya's verified historical payroll data."""

    def _meta(start, end, pay):
        s = date.fromisoformat(start)
        e = date.fromisoformat(end)
        p = date.fromisoformat(pay)
        return {
            'dates':        f"{s.strftime('%b %-d')} \u2013 {e.strftime('%b %-d')}, {e.year}",
            'payDate':      p.strftime('%b %-d, %Y'),
            'period_start': start,
            'period_end':   end,
            'pay_date_iso': pay,
            'state':        'locked',
            'locked':       True,
            'lockedAt':     pay,
            'lockedBy':     'Tanya (historical)',
            'hardcoded':    True,
            'daysElapsed':  (e - s).days + 1,
            'daysTotal':    (e - s).days + 1,
        }

    periods = {}

    # ── 1.16.26 (Dec 27, 2025 – Jan 9, 2026) ──────────────────────────────
    p = _meta('2025-12-27', '2026-01-09', '2026-01-16')
    p['label'] = '1.16.26'
    p['snapshot'] = {
        'doctors': [
            {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 40706.06, 'payNo': 14654.18, 'payWith': 17498.80, 'off': {'MV': 15612.50, 'MO': 1022.52, 'SW': 24071.04}},
            {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 36138.64, 'payNo': 11564.36, 'payWith': 13029.92, 'off': {'MV': 8942.95, 'MO': 19031.63, 'LS': 356.06, 'EV': 7808.00}},
            {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 33875.44, 'payNo': 11856.40, 'payWith': 14055.46, 'off': {'MV': 10324.63, 'SW': 23449.65, 'EV': 101.16}},
            {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 30738.98, 'payNo': 10758.64, 'payWith': 12472.02, 'off': {'MV': 4357.82, 'MO': 20110.41, 'EV': 6270.75}},
            {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 27164.96, 'payNo': 8421.14,  'payWith': 10119.69, 'off': {'MV': 13678.80, 'MO': 13269.16, 'SW': 217.00}},
            {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 21251.67, 'payNo': 7438.08,  'payWith': 8879.72,  'off': {'MV': 670.20, 'MO': 19461.27, 'EV': 1120.20}},
            {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 450.70,   'payNo': 144.22,   'payWith': 175.90,   'off': {'MV': 101.20, 'MO': 349.50}},
            {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 0.00,     'payNo': 0.00,     'payWith': 0.00},
        ],
        'offices': [
            {'name': 'Monroe',       'amt': 82545.87},
            {'name': 'Marysville',   'amt': 58154.50},
            {'name': 'Stanwood',     'amt': 47737.69},
            {'name': 'Everett',      'amt': 30446.41},
            {'name': 'Lake Stevens', 'amt': 356.06},
        ],
        'termed': [{'name': 'Dr. Kirk', 'coll': 28914.08, 'note': 'Terminated'}],
    }
    # For locked periods, snapshot IS the data (no separate live field)
    p['doctors'] = p['snapshot']['doctors']
    p['offices'] = p['snapshot']['offices']
    p['termed']  = p['snapshot']['termed']
    periods['1.16.26'] = p

    # ── 1.30.26 (Jan 10 – Jan 23, 2026) ───────────────────────────────────
    p = _meta('2026-01-10', '2026-01-23', '2026-01-30')
    p['label'] = '1.30.26'
    p['snapshot'] = {
        'doctors': [
            {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 57272.45, 'payNo': 20618.08, 'payWith': 23550.87, 'off': {'MV': 22683.98, 'MO': 7543.50, 'SW': 27044.97}},
            {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 52757.04, 'payNo': 18464.96, 'payWith': 21338.47, 'off': {'MV': 10645.82, 'MO': 30920.14, 'LS': 4514.09, 'EV': 6676.99}},
            {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 46867.49, 'payNo': 14997.60, 'payWith': 17744.43, 'off': {'MV': 21326.45, 'MO': 22960.86, 'LS': 463.40, 'EV': 2116.78}},
            {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 42240.48, 'payNo': 14784.17, 'payWith': 17380.04, 'off': {'MV': 20835.35, 'SW': 21071.26, 'EV': 333.87}},
            {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 29843.79, 'payNo': 9251.57,  'payWith': 10988.61, 'off': {'MV': 16268.65, 'MO': 3955.53, 'LS': 2636.10, 'SW': 56.94, 'EV': 6926.57}},
            {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 21932.74, 'payNo': 7676.46,  'payWith': 8943.45,  'off': {'MV': 527.60, 'MO': 16752.57, 'LS': 1850.75, 'EV': 2801.82}},
            {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 1642.76,  'payNo': 739.24,   'payWith': 739.24,   'off': {'MV': 1491.16, 'EV': 151.60}},
            {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 445.10,   'payNo': 142.43,   'payWith': 147.17,   'off': {'MO': 445.10}},
        ],
        'offices': [
            {'name': 'Marysville',   'amt': 100800.74},
            {'name': 'Monroe',       'amt': 84035.16},
            {'name': 'Stanwood',     'amt': 48842.88},
            {'name': 'Everett',      'amt': 26515.45},
            {'name': 'Lake Stevens', 'amt': 10345.60},
        ],
        'termed': [{'name': 'Dr. Kirk', 'coll': 17537.98, 'note': 'Terminated'}],
    }
    p['doctors'] = p['snapshot']['doctors']
    p['offices'] = p['snapshot']['offices']
    p['termed']  = p['snapshot']['termed']
    periods['1.30.26'] = p

    # ── 2.13.26 (Jan 24 – Feb 6, 2026) — Benton rate = 0.33 ──────────────
    p = _meta('2026-01-24', '2026-02-06', '2026-02-13')
    p['label'] = '2.13.26'
    p['snapshot'] = {
        'doctors': [
            {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 70435.18, 'payNo': 25356.66, 'payWith': 29323.52, 'off': {'MV': 34056.54, 'MO': 2468.80, 'SW': 33909.84}},
            {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 54346.07, 'payNo': 16847.28, 'payWith': 18518.56, 'off': {'MV': 9983.65, 'MO': 5563.26, 'LS': 22371.29, 'SW': 937.98, 'EV': 15489.89}},
            {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 52215.52, 'payNo': 18275.43, 'payWith': 21036.47, 'off': {'MV': 23695.69, 'MO': 3615.50, 'SW': 24904.33}},
            {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 43717.91, 'payNo': 15301.27, 'payWith': 17589.10, 'off': {'MV': 268.32, 'MO': 24897.54, 'LS': 12179.74, 'EV': 6372.31}},
            {'name': 'Dr. Benton',  'pct': 0.33, 'coll': 38867.69, 'payNo': 12826.34, 'payWith': 14665.29, 'off': {'MV': 7224.62, 'MO': 24637.62, 'LS': 6652.45, 'EV': 353.00}},
            {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 24586.75, 'payNo': 8605.36,  'payWith': 9617.71,  'off': {'MV': 258.60, 'MO': 12073.41, 'LS': 12041.77, 'EV': 212.97}},
            {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 5443.12,  'payNo': 1741.80,  'payWith': 2141.82,  'off': {'MV': 1376.24, 'MO': 4066.88}},
            {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 2023.41,  'payNo': 910.53,   'payWith': 910.53,   'off': {'MV': 1544.01, 'EV': 479.40}},
        ],
        'offices': [
            {'name': 'Marysville',   'amt': 78477.67},
            {'name': 'Monroe',       'amt': 77629.78},
            {'name': 'Stanwood',     'amt': 59796.15},
            {'name': 'Lake Stevens', 'amt': 53245.25},
            {'name': 'Everett',      'amt': 23858.82},
        ],
        'termed': [{'name': 'Dr. Kirk', 'coll': 1372.02, 'note': 'Terminated'}],
    }
    p['doctors'] = p['snapshot']['doctors']
    p['offices'] = p['snapshot']['offices']
    p['termed']  = p['snapshot']['termed']
    periods['2.13.26'] = p

    # ── 2.27.26 (Feb 7 – Feb 20, 2026) — Benton rate = 0.32 ──────────────
    p = _meta('2026-02-07', '2026-02-20', '2026-02-27')
    p['label'] = '2.27.26'
    p['snapshot'] = {
        'doctors': [
            {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 68512.61, 'payNo': 24664.54, 'payWith': 28572.15, 'off': {'MV': 27086.51, 'MO': 11879.86, 'SW': 29546.24}},
            {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 49682.98, 'payNo': 17389.04, 'payWith': 19276.69, 'off': {'MV': 157.00, 'MO': 20429.10, 'LS': 14883.05, 'EV': 14213.83}},
            {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 47354.85, 'payNo': 16574.20, 'payWith': 18910.47, 'off': {'MV': 19626.25, 'MO': 1889.00, 'SW': 25839.60}},
            {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 43720.53, 'payNo': 13990.57, 'payWith': 15844.07, 'off': {'MV': 12134.16, 'MO': 18453.36, 'LS': 13133.01}},
            {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 38646.25, 'payNo': 11980.34, 'payWith': 13737.56, 'off': {'MV': 5937.86, 'MO': 168.96, 'LS': 20983.32, 'SW': 4166.44, 'EV': 7389.67}},
            {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 31620.97, 'payNo': 10118.71, 'payWith': 11718.98, 'off': {'MV': 7235.48, 'MO': 24385.49}},
            {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 25535.53, 'payNo': 8937.44,  'payWith': 10114.36, 'off': {'MV': 41.80, 'MO': 10481.83, 'LS': 15011.90}},
            {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 7528.70,  'payNo': 3387.91,  'payWith': 3387.91,  'off': {'MV': 4113.00, 'EV': 3415.70}},
        ],
        'offices': [
            {'name': 'Monroe',       'amt': 87687.60},
            {'name': 'Marysville',   'amt': 76332.06},
            {'name': 'Lake Stevens', 'amt': 64011.28},
            {'name': 'Stanwood',     'amt': 59552.28},
            {'name': 'Everett',      'amt': 25019.20},
        ],
        'termed': [{'name': 'Dr. Kirk', 'coll': 0.00, 'note': 'Terminated'}],
    }
    p['doctors'] = p['snapshot']['doctors']
    p['offices'] = p['snapshot']['offices']
    p['termed']  = p['snapshot']['termed']
    periods['2.27.26'] = p

    # ── 3.13.26 (Feb 21 – Mar 6, 2026) — first SQL-derived locked period ──
    p = _meta('2026-02-21', '2026-03-06', '2026-03-13')
    p['label'] = '3.13.26'
    p['snapshot'] = {
        'doctors': [
            {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 70370.80, 'payNo': 22014.26, 'payWith': 25333.49, 'off': {'SW': 25837.06, 'MV': 36801.98, 'MO': 7331.76, 'EV': 400.00}},
            {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 58956.71, 'payNo': 17920.60, 'payWith': 20634.85, 'off': {'SW': 31712.29, 'MV': 27198.71, 'MO': 45.71}},
            {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 58330.50, 'payNo': 16710.01, 'payWith': 18665.76, 'off': {'MO': 35229.01, 'MV': 8993.60, 'LS': 13580.89, 'EV': 527.00}},
            {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 41515.69, 'payNo': 11080.56, 'payWith': 12869.86, 'off': {'MV': 9800.89, 'MO': 5244.14, 'LS': 19080.06, 'EV': 7390.60}},
            {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 38689.60, 'payNo': 11751.30, 'payWith': 13541.36, 'off': {'MO': 20103.58, 'LS': 12984.08, 'EV': 5556.94, 'MV': 45.00}},
            {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 23163.76, 'payNo': 6412.57,  'payWith': 7412.40,  'off': {'MO': 18271.04, 'MV': 4526.72, 'EV': 366.00}},
            {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 23154.88, 'payNo': 7147.49,  'payWith': 8104.21,  'off': {'MV': 200.00, 'MO': 8228.25, 'EV': 6.00, 'LS': 14720.63}},
            {'name': 'Dr. Lugo',    'pct': 0.36, 'coll': 13610.13, 'payNo': 0.00,     'payWith': 4899.65,  'off': {'MV': 11514.18, 'MO': 535.00, 'SW': 3.82, 'LS': 71.13, 'EV': 1486.00}},
            {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 5030.47,  'payNo': 2263.71,  'payWith': 2263.71,  'off': {'EV': 659.00, 'MV': 4371.47}},
        ],
        'offices': [
            {'name': 'Marysville',   'amt': 103452.55},
            {'name': 'Monroe',       'amt': 95325.59},
            {'name': 'Lake Stevens', 'amt': 60467.79},
            {'name': 'Stanwood',     'amt': 57859.23},
            {'name': 'Everett',      'amt': 16821.54},
        ],
        'termed': [
            {'name': 'Dr. Kirk',  'coll': 31.00,   'note': 'Account closed'},
            {'name': 'Dr. Ping',  'coll': 817.10,  'note': 'Terminated'},
            {'name': 'Dr. Laws',  'coll': 0.00,    'note': 'Terminated'},
        ],
    }
    p['doctors'] = p['snapshot']['doctors']
    p['offices'] = p['snapshot']['offices']
    p['termed']  = p['snapshot']['termed']
    periods['3.13.26'] = p

    return periods


# ─────────────────────────────────────────────────────────────────────────────
# THREE-STATE LIFECYCLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def get_period_state(period_end_obj, pay_date_obj):
    """
    Compute the three-phase state from dates alone.
    Callers should check for 'locked' state in the JSON before using this.
    NOTE: live → pending transition is triggered by Sunday 5am CST run,
    not just by date. This function provides a date-based fallback only.
    """
    today = date.today()
    if today <= period_end_obj:
        return 'live'
    elif today < pay_date_obj:
        return 'pending'
    else:
        return 'locked'


def is_period_locked(period_data):
    """Return True if period is permanently locked (human confirmed)."""
    return period_data.get('state') == 'locked' or period_data.get('locked', False)


def should_transition_to_pending(period_data):
    """
    Returns True if we should flip this period from live → pending.

    Rules:
    - Period must currently be 'live' in JSON
    - Period end date must be in the past (today > period_end)
    - Script must be running on a Sunday at 5am CST/CDT (Azure task schedule)
      OR we are past the Sunday-after-period-end (safety catch for missed runs)

    Sunday at 5am CST = the Denticon Saturday morning dump has settled.
    A pending period is COMPLETELY FROZEN after this snapshot — zero updates.

    TIME GUARD (UTC):
    ─────────────────
    5am CST  = 11:00 UTC  (UTC-6, standard time Nov–Mar)
    5am CDT  = 10:00 UTC  (UTC-5, daylight time Mar–Nov)

    DST in the US starts the second Sunday in March.  For any period ending
    in February or later (Mar 8+ in 2026), the snapshot Sunday is in CDT.
    We use 10:00 UTC (5am CDT) as the cutoff to be accurate for most of the
    year.  We also accept ≥ 11:00 UTC to cover the rare CST case and manual
    re-runs later in the day.

    To prevent an accidental early snapshot (e.g. manual run at 2am CDT
    on Sunday before Azure has had a chance to fully dump Denticon), we
    require utcnow().hour >= 10 on the snapshot Sunday itself.
    For days AFTER the snapshot Sunday (missed-run safety catch) there is
    no time restriction.
    """
    state = period_data.get('state', period_data.get('status', 'live'))
    if state != 'live':
        return False
    period_end_str = period_data.get('period_end', '')
    if not period_end_str:
        return False
    try:
        period_end = date.fromisoformat(period_end_str)
        today = date.today()
        if today <= period_end:
            return False  # Period still active

        # Find the Sunday after period_end
        days_until_sunday = (6 - period_end.weekday()) % 7  # 6 = Sunday
        if days_until_sunday == 0:
            days_until_sunday = 7  # If period ended on Sunday, use next Sunday
        snapshot_sunday = period_end + timedelta(days=days_until_sunday)

        if today < snapshot_sunday:
            return False  # Haven't reached snapshot Sunday yet

        if today == snapshot_sunday:
            # Enforce the 5am CDT (10:00 UTC) time guard on the snapshot day itself.
            # This prevents accidental early snapshots on manual runs before the
            # Denticon dump has settled.
            utc_hour = datetime.utcnow().hour
            if utc_hour < 10:
                print(f"    INFO: Snapshot Sunday reached but UTC hour={utc_hour} < 10 "
                      f"(5am CDT). Waiting for full Azure dump. Skipping transition.")
                return False

        return True
    except ValueError:
        return False


def make_snapshot(period_data):
    """
    Copy live doctor/office/termed arrays into a snapshot dict.
    Called when transitioning live → processing.
    Returns snapshot dict (immutable reference copy of current data).
    """
    return {
        'doctors':   [dict(d) for d in period_data.get('doctors', [])],
        'offices':   [dict(o) for o in period_data.get('offices', [])],
        'termed':    [dict(t) for t in period_data.get('termed',  [])],
        'snapshotAt': datetime.utcnow().isoformat() + 'Z',
    }


# ─────────────────────────────────────────────────────────────────────────────
# AZURE SQL CONNECTION
# ─────────────────────────────────────────────────────────────────────────────

def get_connection(conn_str):
    """Create Azure SQL connection."""
    try:
        import pyodbc
    except ImportError:
        print("ERROR: pyodbc not installed. Run: pip install pyodbc")
        sys.exit(1)

    drivers = pyodbc.drivers()
    print(f"  Available ODBC drivers: {drivers}")

    cs_lower = conn_str.lower()
    if 'your-' in cs_lower or 'your_' in cs_lower:
        print("ERROR: Connection string contains placeholder values")
        sys.exit(1)
    if 'driver' not in cs_lower:
        print("WARNING: Connection string missing 'Driver=' parameter")

    try:
        conn = pyodbc.connect(conn_str, timeout=30)
        print("  Connected to Azure SQL")
        return conn
    except Exception as e:
        print(f"ERROR: Azure SQL connection failed: {e}")
        import re as _re
        m = _re.search(r'Server\s*=\s*tcp:([^,;]+)', conn_str, _re.IGNORECASE)
        if m:
            print(f"  Server: {m.group(1)}")
        sys.exit(1)


def load_id_maps(conn):
    """Load provider and office ID→name maps from Denticon tables."""
    cursor = conn.cursor()
    cursor.execute("SELECT PROVIDERID, LNAME, FNAME FROM PGID4951_PROVIDER WHERE LNAME IS NOT NULL")
    for r in cursor.fetchall():
        PROVIDER_ID_MAP[r.PROVIDERID] = f"{r.LNAME.strip()}, {r.FNAME.strip()}"
    cursor.execute("SELECT OID, OFFICENAME FROM PGID4951_OFFICE")
    for r in cursor.fetchall():
        OFFICE_ID_MAP[r.OID] = r.OFFICENAME.strip() if r.OFFICENAME else f"Office {r.OID}"
    print(f"  ID maps: {len(PROVIDER_ID_MAP)} providers, {len(OFFICE_ID_MAP)} offices")


def get_pay_periods(conn):
    """Fetch pay period definitions from rpt.vw_pay_periods."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pay_period_num, period_start, period_end
        FROM rpt.vw_pay_periods
        ORDER BY period_start
    """)
    periods = []
    for r in cursor.fetchall():
        s = r.period_start if isinstance(r.period_start, date) else date.fromisoformat(str(r.period_start)[:10])
        e = r.period_end   if isinstance(r.period_end,   date) else date.fromisoformat(str(r.period_end)[:10])
        pay = e + timedelta(days=7)
        label = f"{pay.month}.{pay.day}.{str(pay.year)[2:]}"
        periods.append({'num': r.pay_period_num, 'start': s, 'end': e, 'pay_date': pay, 'label': label})
    return periods


def find_current_period(periods):
    """Find the period containing today, or the most recent past period."""
    today = date.today()
    for p in periods:
        if p['start'] <= today <= p['end']:
            return p
    past = [p for p in periods if p['end'] < today]
    return past[-1] if past else (periods[0] if periods else None)


# ─────────────────────────────────────────────────────────────────────────────
# TRANSACTION QUERY — uses rpt.vw_income_allocation (Tanya's report)
# ─────────────────────────────────────────────────────────────────────────────

def query_income_allocation(conn, period_start, period_end):
    """
    Query rpt.vw_income_allocation for the given period.

    DATA SOURCE NOTE:
    ─────────────────
    rpt.vw_income_allocation is the Azure SQL mirror of Denticon's
    "Income Allocation Report - Detail" — the same report Tanya runs manually.
    This is the AUTHORITATIVE source for pay calculations.

    We do NOT use rpt.vw_doctor_collections_no_xray because:
    1. We need transaction-level proc_ada_code to apply the x-ray exclusion
       ourselves per doctor (Schrack is exempt; owner is exempt).
    2. Pre-aggregated views may round at aggregation time, causing the ~0.02%
       variance seen against Tanya's spreadsheet.
    3. Transaction-level data lets us audit individual procedures if needed.

    SIGN CONVENTION:
    ────────────────
    Denticon income values are SIGNED:
      - Payments received = negative
      - Adjustments       = positive
    We sum raw signed values per provider, then take abs() of the NET total.
    This matches how Tanya's report displays values (abs of net sum).
    """
    cursor = conn.cursor()
    # Filter by ALLOCDATE (the actual transaction date) between period boundaries.
    # This is the same date range Tanya's Income Allocation Report uses when she
    # runs it for a specific period — it is the correct field to use.
    #
    # We also include period_start/period_end view metadata columns if the view
    # supports them; if not, the ALLOCDATE filter alone is the source of truth.
    #
    # NOTE: Do NOT filter by proc_date — alloc_date is the posted/allocated date
    # that Denticon uses for period assignment and that Tanya's report uses.
    # Any same-day postings after the 1am Saturday dump will appear in the NEXT
    # period (expected ~$21K gap for today's late postings — this is correct).
    cursor.execute("""
        SELECT
            alloc_provider_id,
            OID            AS office_id,
            ALLOCDATE,
            alloc_amount,
            proc_ada_code
        FROM rpt.vw_income_allocation
        WHERE ALLOCDATE BETWEEN ? AND ?
    """, (str(period_start), str(period_end)))

    transactions = []
    for r in cursor.fetchall():
        provider = PROVIDER_ID_MAP.get(r.alloc_provider_id, f"Provider {r.alloc_provider_id}")
        office   = OFFICE_ID_MAP.get(r.office_id, f"Office {r.office_id}")
        office_key = office.upper().replace('PSPD - ', '').replace('PSPD-', '').strip()
        transactions.append({
            'office':     office_key,
            'provider':   provider,
            'alloc_date': r.ALLOCDATE,
            'proc_code':  str(r.proc_ada_code).strip() if r.proc_ada_code else '',
            'income':     float(r.alloc_amount) if r.alloc_amount else 0.0,
        })

    print(f"    rpt.vw_income_allocation: {len(transactions)} rows")
    return transactions


# ─────────────────────────────────────────────────────────────────────────────
# X-RAY CODE AUDIT
# ─────────────────────────────────────────────────────────────────────────────
#
# Background: Non-exempt doctors are paid on (collections - xray_collections).
# We currently exclude codes starting with D02* and D03* as x-ray procedures.
# If that prefix list is wrong, some codes may be wrongly excluded (inflating
# payNo reduction) or wrongly included (missing x-ray codes, inflating payNo).
#
# Observed discrepancies vs Tanya's manual Denticon report (Mar 2026):
#   Slaven:  dashboard ~$88  LESS  → x-ray over-excluded (extra D02/D03 non-xray?)
#   Choong:  dashboard ~$111 MORE  → x-ray under-excluded (missing xray code?)
#             OR a D02/D03 code that is NOT xray is being wrongly excluded for her
#   Welter:  dashboard ~$30  LESS  → x-ray over-excluded
#   Menon:   dashboard ~$52  LESS  → x-ray over-excluded
#   Patel:   dashboard ~$48  LESS  → x-ray over-excluded
#   Bell:    PERFECT match
#   Schrack: PERFECT match (xray_exempt=True, exclusion doesn't apply)

# Known x-ray ADA codes (for audit cross-reference):
# D0210 - Full mouth series (FMX)
# D0220 - Periapical x-ray, first image
# D0230 - Periapical x-ray, each additional
# D0240 - Occlusal radiographic image
# D0270 - Bitewing, single image
# D0272 - Bitewings, two images
# D0273 - Bitewings, three images
# D0274 - Bitewings, four images
# D0277 - Vertical bitewings, 7-8 images
# D0290 - Posterior-anterior or lateral skull
# D0310 - Sialography
# D0320 - Temporomandibular joint arthrogram
# D0322 - Tomographic survey
# D0330 - Panoramic radiographic image
# D0340 - 2-D cephalometric radiographic image
# D0350 - 2-D oral/facial photographic image (NOT an x-ray!)
# D0360 - Cone beam CT (CBCT), single arch
# D0363 - Cone beam CT, both arches
# D0364-D0368 - CBCT variations
# D0380-D0386 - CBCT additional views
# D0391 - Interpretation of x-rays (NOT an x-ray image itself)
# D0394 - Digital subtraction
# D0395 - Facial photo (NOT an x-ray!)

# Codes that start with D02/D03 but are NOT x-ray imaging:
NON_XRAY_D02_D03_CODES = {
    'D0350',  # 2-D oral/facial photographic image (photo, not radiation)
    'D0391',  # Interpretation of radiographs (interpretive service, not imaging)
    'D0393',  # Treatment simulation using 3D image volume
    'D0394',  # Digital subtraction of two or more images
    'D0395',  # Facial photo image obtained to interpret or compare
}

# X-ray codes that do NOT start with D02 or D03 (would be missed by prefix filter):
XRAY_CODES_OUTSIDE_PREFIX = {
    # Currently none identified in ADA CDT — D02/D03 covers all standard
    # diagnostic radiology. If Denticon uses custom/local codes (e.g. D9999
    # or non-standard), they would appear here in the audit output.
}


def audit_xray_codes(transactions):
    """
    Audit procedure codes in the transaction set to identify:
    1. All distinct D0* codes present in the data
    2. Any x-ray codes outside D02*/D03* prefixes (missed by current filter)
    3. Any D02*/D03* codes that are NOT diagnostic x-ray images (wrongly excluded)
    4. Per-doctor breakdown of what's being excluded (to debug directional gaps)

    This function DOES NOT change any pay calculations — audit only.
    Call with the raw transaction list from query_income_allocation().
    """
    print("\n─── X-Ray Code Audit ────────────────────────────────────────")

    # 1. Collect all distinct D0* codes
    d0_codes = {}   # code → {doctors: set, count: int, total_income: float}
    for t in transactions:
        code = t.get('proc_code', '').strip()
        if code.upper().startswith('D0'):
            key = code.upper()
            if key not in d0_codes:
                d0_codes[key] = {'doctors': set(), 'count': 0, 'total_income': 0.0}
            d0_codes[key]['doctors'].add(t.get('provider', '?'))
            d0_codes[key]['count'] += 1
            d0_codes[key]['total_income'] += t.get('income', 0.0)

    print(f"\n  1. Distinct D0* codes in dataset ({len(d0_codes)} total):")
    for code in sorted(d0_codes.keys()):
        info = d0_codes[code]
        is_excluded = code.startswith(XRAY_PREFIXES)
        is_known_non_xray = code in NON_XRAY_D02_D03_CODES
        flag = ''
        if is_excluded and is_known_non_xray:
            flag = '  ⚠️  WRONGLY EXCLUDED (D02/D03 but NOT an x-ray)'
        elif is_excluded:
            flag = '  ✓ excluded as x-ray'
        else:
            flag = '  — not excluded'
        print(f"    {code:8s}  count={info['count']:4d}  net=${abs(info['total_income']):>10,.2f}  "
              f"doctors={len(info['doctors'])}{flag}")

    # 2. Check for x-ray codes outside D02*/D03* (missed by prefix filter)
    print(f"\n  2. X-ray codes outside D02*/D03* prefix (would be MISSED):")
    missed = {c: d0_codes[c] for c in d0_codes if c in XRAY_CODES_OUTSIDE_PREFIX}
    if missed:
        for code, info in missed.items():
            print(f"    ⚠️  {code}  count={info['count']}  net=${abs(info['total_income']):,.2f}")
    else:
        print("    None found — all known x-ray codes are within D02*/D03*")
    # Also flag any non-D02/D03 codes that look x-ray-ish (heuristic)
    suspicious_non_prefix = []
    for code, info in d0_codes.items():
        if not code.startswith(XRAY_PREFIXES) and code not in XRAY_CODES_OUTSIDE_PREFIX:
            # These would be missed by current filter — flag for human review
            suspicious_non_prefix.append((code, info))
    if suspicious_non_prefix:
        print(f"    D0* codes NOT in D02/D03 (verify these aren't x-rays):")
        for code, info in sorted(suspicious_non_prefix):
            print(f"      {code:8s}  count={info['count']:4d}  net=${abs(info['total_income']):>10,.2f}")

    # 3. Check for D02*/D03* codes that are NOT x-ray procedures
    print(f"\n  3. D02*/D03* codes that are NOT diagnostic x-ray imaging (wrongly excluded):")
    wrongly_excluded = {c: d0_codes[c] for c in d0_codes
                        if c.startswith(XRAY_PREFIXES) and c in NON_XRAY_D02_D03_CODES}
    if wrongly_excluded:
        for code, info in wrongly_excluded.items():
            print(f"    ⚠️  {code}  count={info['count']}  net=${abs(info['total_income']):,.2f}"
                  f"  doctors={info['doctors']}")
            print(f"       → Should NOT be excluded from payroll base")
    else:
        print("    None found — all D02/D03 codes in dataset appear to be genuine x-rays")

    # 4. Per-doctor x-ray exclusion breakdown (to debug directional gaps)
    print(f"\n  4. Per-doctor x-ray exclusion breakdown:")
    doc_xray = {}  # provider → {xray_total, non_xray_d0_total, all_total}
    for t in transactions:
        prov = t.get('provider', '?')
        code = t.get('proc_code', '').strip().upper()
        income = t.get('income', 0.0)
        if prov not in doc_xray:
            doc_xray[prov] = {'xray': 0.0, 'wrongly_excluded': 0.0, 'total': 0.0}
        doc_xray[prov]['total'] += income
        if code.startswith(XRAY_PREFIXES):
            doc_xray[prov]['xray'] += income
            if code in NON_XRAY_D02_D03_CODES:
                doc_xray[prov]['wrongly_excluded'] += income

    for prov in sorted(doc_xray.keys()):
        info = doc_xray[prov]
        xray_abs = abs(info['xray'])
        wrong_abs = abs(info['wrongly_excluded'])
        if xray_abs > 0.01:
            note = f'  ⚠️  incl. ${wrong_abs:,.2f} wrongly excluded' if wrong_abs > 0.01 else ''
            print(f"    {prov:30s}  x-ray excluded=${xray_abs:>8,.2f}{note}")
        elif abs(info['total']) > 0.01:
            print(f"    {prov:30s}  x-ray excluded=       $0.00  (no x-ray codes)")

    print("─── End X-Ray Audit ─────────────────────────────────────────\n")


# ─────────────────────────────────────────────────────────────────────────────
# PAY CALCULATION
# ─────────────────────────────────────────────────────────────────────────────

def process_transactions(transactions, period_start, period_end, label, pay_date):
    """
    Transform raw income allocation transactions into payroll.json period format.

    PAY MATH RULES:
    ───────────────
    Standard associate:
        payWith = coll × rate
        payNo   = (coll - xray) × rate

    X-ray-exempt (Dr. Schrack, Dr. Lugo/owner):
        payWith = coll × rate
        payNo   = 0.0          ← Dr. Lugo convention (savings shown separately)
                = coll × rate  ← Dr. Schrack (payNo = payWith, no exclusion)

    ROUNDING RULE: accumulate in full-precision float. Round ONLY at final
    output. Do NOT round coll or xray before computing pay — that causes
    the 0.02% gap vs Tanya's spreadsheet.
    """
    today = date.today()
    days_total   = (period_end - period_start).days + 1
    days_elapsed = max(1, min(days_total, (today - period_start).days + 1))

    # ── Aggregate raw signed values per provider + office ──────────────────
    # Use full-precision accumulation (no intermediate rounding)
    raw_prov  = {}   # provider_name → {total, xray, offices: {office_key: total}}
    raw_office = {}  # office_key → total

    for t in transactions:
        prov      = t['provider']
        office    = t['office']
        income    = t['income']   # signed float
        proc_code = t['proc_code']
        is_xray   = proc_code.startswith(XRAY_PREFIXES)

        if prov not in raw_prov:
            raw_prov[prov] = {'total': 0.0, 'xray': 0.0, 'offices': {}}
        raw_prov[prov]['total'] += income
        if is_xray:
            raw_prov[prov]['xray'] += income
        raw_prov[prov]['offices'].setdefault(office, 0.0)
        raw_prov[prov]['offices'][office] += income

        raw_office.setdefault(office, 0.0)
        raw_office[office] += income

    # ── Build doctor array (active doctors only) ───────────────────────────
    doctors = []
    for denticon_name, config in DOCTOR_CONFIG.items():
        pdata = raw_prov.get(denticon_name, {'total': 0.0, 'xray': 0.0, 'offices': {}})

        # Full-precision absolute values (sign convention: collections are negative in DB)
        coll_fp = abs(pdata['total'])
        xray_fp = abs(pdata['xray'])
        rate    = config['pct']
        is_owner       = config.get('owner', False)
        is_xray_exempt = config.get('xray_exempt', False)

        # Compute pay in full precision, round only at the end
        pay_with_fp = coll_fp * rate
        if is_owner:
            # Owner: show only payWith; payNo = 0 by convention
            pay_no_fp = 0.0
        elif is_xray_exempt:
            # Schrack (1099 oral surgeon): no x-ray exclusion applies
            # payNo = payWith (he doesn't do diagnostic x-rays)
            pay_no_fp = pay_with_fp
        else:
            # Standard associate: exclude x-ray collections from pay base
            pay_no_fp = (coll_fp - xray_fp) * rate

        # Round only here — final output step
        coll     = round(coll_fp,     2)
        xray     = round(xray_fp,     2)
        pay_with = round(pay_with_fp, 2)
        pay_no   = round(pay_no_fp,   2)

        doc_entry = {
            'name':    config['display'],
            'pct':     rate,
            'coll':    coll,
            'payNo':   pay_no,
            'payWith': pay_with,
        }

        # Per-office breakdown
        off = {}
        for raw_off_name, amt in pdata['offices'].items():
            abs_amt = abs(amt)
            if abs_amt < 0.01:
                continue
            for office_key, (_, abbr) in OFFICE_NAME_MAP.items():
                if office_key in raw_off_name.upper():
                    off[abbr] = round(abs_amt, 2)
                    break
        if off:
            doc_entry['off'] = off

        doctors.append(doc_entry)

    # Sort by collections descending
    doctors.sort(key=lambda d: d['coll'], reverse=True)

    # ── Office totals ──────────────────────────────────────────────────────
    offices = []
    for raw_off_name, total in sorted(raw_office.items(), key=lambda x: -abs(x[1])):
        abs_total = abs(total)
        if abs_total < 0.01:
            continue
        display = raw_off_name
        for office_key, (disp, _) in OFFICE_NAME_MAP.items():
            if office_key in raw_off_name.upper():
                display = disp
                break
        offices.append({'name': display, 'amt': round(abs_total, 2)})

    # ── Termed doctors (residual collections) ─────────────────────────────
    termed = []
    for denticon_name, config in TERMED_DOCTORS.items():
        pdata = raw_prov.get(denticon_name, {'total': 0.0})
        termed.append({
            'name': config['display'],
            'coll': round(abs(pdata['total']), 2),
            'note': config['note'],
        })

    def fmt_short(d):
        return d.strftime('%b %-d')

    period = {
        'label':        label,
        'dates':        f"{fmt_short(period_start)} \u2013 {fmt_short(period_end)}, {period_end.year}",
        'payDate':      pay_date.strftime('%b %-d, %Y'),
        'period_start': str(period_start),
        'period_end':   str(period_end),
        'pay_date_iso': str(pay_date),
        'daysElapsed':  days_elapsed,
        'daysTotal':    days_total,
        'doctors':      doctors,
        'offices':      offices,
        'termed':       termed,
    }

    return period


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL PARSING — fallback for Tanya's manual export
# ─────────────────────────────────────────────────────────────────────────────

def parse_excel(filepath):
    """Parse Denticon 'Income Allocation Report - Detail' Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    transactions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        office = None
        current_provider = None

        for row in ws.iter_rows(min_row=1, values_only=False):
            a_val = row[0].value if len(row) > 0 else None
            b_val = row[1].value if len(row) > 1 else None
            c_val = row[2].value if len(row) > 2 else None

            if a_val and isinstance(a_val, str) and 'Office:' in a_val:
                m = re.search(r'PSPD\s*-\s*(.+)', a_val)
                if m:
                    office = m.group(1).strip().upper()
                continue

            if b_val and isinstance(b_val, str) and 'Provider :-' in b_val:
                parts = b_val.split(':-')[1].strip()
                m = re.match(r'([^,]+,\s*\S+)', parts)
                if m:
                    current_provider = m.group(1).strip()
                continue

            # Skip Grand Total / summary rows: they have a string in column A
            # (e.g. "Grand Total", "Provider Total") rather than a date in column C.
            # Only transaction rows have a datetime in column C.
            if a_val and isinstance(a_val, str) and any(
                kw in a_val for kw in ('Grand Total', 'Provider Total', 'Office Total', 'Report Total')
            ):
                continue

            if c_val and isinstance(c_val, datetime) and current_provider and office:
                income    = row[14].value if len(row) > 14 else None
                proc_code = row[12].value if len(row) > 12 else None
                if income is not None:
                    transactions.append({
                        'office':     office,
                        'provider':   current_provider,
                        'alloc_date': c_val,
                        'proc_code':  str(proc_code).strip() if proc_code else '',
                        'income':     float(income),
                    })

    wb.close()
    print(f"  Parsed {len(transactions)} transactions from {Path(filepath).name}")
    return transactions


def find_period_for_transactions(transactions, override_label=None):
    """Detect which pay period the Excel transactions fall in."""
    if not transactions:
        return None

    if override_label:
        for label, start, end, pay in FALLBACK_PAY_PERIODS:
            if label == override_label:
                return label, date.fromisoformat(start), date.fromisoformat(end), date.fromisoformat(pay)
        return None

    dates = [t['alloc_date'] for t in transactions if isinstance(t['alloc_date'], datetime)]
    if not dates:
        return None
    min_d = min(dates).date()
    max_d = max(dates).date()

    for label, start, end, pay in FALLBACK_PAY_PERIODS:
        s = date.fromisoformat(start)
        e = date.fromisoformat(end)
        if min_d <= e and max_d >= s:
            return label, s, e, date.fromisoformat(pay)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def load_existing_json(output_path):
    """Load existing payroll.json. Returns dict of periods keyed by label."""
    try:
        with open(output_path, 'r') as f:
            data = json.load(f)
        return data.get('periods', {})
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def run_azure_pipeline(conn_str, target, output_path, audit_xray=False):
    """
    Full pipeline with three-state period management:

    1. Load existing payroll.json
    2. Inject hardcoded historical periods (always win; never overwritten)
    3. For each SQL-derived period:
       a. LOCKED  → skip entirely (immutable)
       b. PROCESSING and period not ended → update 'live' field only (for drift)
       c. LIVE + period ended → auto-snapshot + transition to processing
       d. LIVE + period active → query fresh data, update as live
    4. Write payroll.json
    """
    print("\n─── Payroll Refresh Pipeline ───────────────────────────────")

    # Step 1: Load hardcoded historical data (immutable)
    hardcoded = get_hardcoded_periods()
    print(f"  Historical periods: {', '.join(sorted(hardcoded.keys()))}")

    # Step 2: Load existing JSON to preserve locked and processing state
    existing = load_existing_json(output_path)
    n_locked  = sum(1 for p in existing.values() if is_period_locked(p))
    n_pending = sum(1 for p in existing.values() if p.get('state') in ('pending', 'processing'))
    print(f"  Existing JSON: {len(existing)} periods ({n_locked} locked, {n_pending} pending)")

    # Step 3: Connect to SQL
    conn = get_connection(conn_str)
    print("\n  Loading ID maps...")
    load_id_maps(conn)

    print("  Fetching pay periods...")
    all_periods = get_pay_periods(conn)
    print(f"  Found {len(all_periods)} pay periods in DB")

    # Determine target periods
    today = date.today()
    if target == 'current':
        cur = find_current_period(all_periods)
        periods_to_process = [cur] if cur else []
    elif target == 'all':
        periods_to_process = [p for p in all_periods if p['start'] <= today]
    elif target == 'recent':
        past = [p for p in all_periods if p['start'] <= today]
        periods_to_process = past[-5:] if len(past) >= 5 else past
    else:
        # Specific label
        match = [p for p in all_periods if p['label'] == target]
        if not match:
            print(f"ERROR: Period '{target}' not found. Available: {', '.join(p['label'] for p in all_periods[:15])}")
            sys.exit(1)
        periods_to_process = match

    # Step 4: Build output dict — start with all locked periods from existing JSON
    periods_out = {}

    # Preserve all locked periods from existing JSON first
    for label, pdata in existing.items():
        if is_period_locked(pdata):
            periods_out[label] = pdata

    # Inject hardcoded periods (always override — these are Tanya's verified numbers)
    for label, hdata in hardcoded.items():
        periods_out[label] = hdata
        hc_str = " *" if hdata.get('hardcoded') else ""
        print(f"  Injected hardcoded{hc_str}: {label}")

    # Step 5: Process each SQL-derived period
    for period_info in periods_to_process:
        label    = period_info['label']
        start    = period_info['start']
        end      = period_info['end']
        pay_date = period_info['pay_date']

        # Never touch locked or hardcoded periods
        if label in periods_out and is_period_locked(periods_out[label]):
            hc = " (hardcoded)" if periods_out[label].get('hardcoded') else ""
            print(f"\n  Period {label}: LOCKED{hc} — skipping")
            continue

        print(f"\n  Period {label} ({start} to {end})...")

        # Determine natural state from dates
        natural_state = get_period_state(end, pay_date)

        # Check existing state in JSON
        existing_period = existing.get(label, {})
        existing_state  = existing_period.get('state', existing_period.get('status', 'live'))

        # ── PENDING: period ended + snapshot taken — completely frozen ───────
        # NEVER update a pending period. It is immutable until locked.
        if existing_state in ('pending', 'processing'):
            snap_date = existing_period.get('snapshotDate') or \
                        (existing_period.get('snapshot') or {}).get('snapshotAt', '?')
            print(f"    State: PENDING (snapshot frozen at {snap_date}) — SKIPPING (no updates)")
            periods_out[label] = existing_period
            continue

        # ── LIVE but period has ended + Sunday has passed → transition to PENDING ──
        if should_transition_to_pending(existing_period) or \
           (natural_state in ('pending', 'locked') and existing_state == 'live'):
            print(f"    Period ended {end} → transitioning live → PENDING (Sunday 5am CST snapshot)")
            print(f"    Querying final data for snapshot...")
            transactions = query_income_allocation(conn, start, end)

            if not transactions:
                print(f"    WARNING: No data for {label} — cannot snapshot, leaving as live")
                continue

            fresh = process_transactions(transactions, start, end, label, pay_date)
            snap  = make_snapshot(fresh)

            # Build the reconciliation note for Tanya — tells her exactly what
            # data source backs the Pending snapshot so she can verify it matches
            # her Monday Income Allocation report run.
            snap_utc = datetime.utcnow()
            # Convert UTC→CST/CDT for human-readable note (CDT = UTC-5 in March)
            snap_cdt = snap_utc - timedelta(hours=5)
            snap_sunday_str  = snap_cdt.strftime('%a %b %-d')  # e.g. "Sun Mar 22"
            snap_sunday_time = snap_cdt.strftime('%-I%p CDT').lower()  # e.g. "5am CDT"
            # Denticon dump is always the prior Saturday at ~1am CST
            denticon_dump_saturday = end + timedelta(days=1)  # Saturday after period_end Friday
            denticon_dump_str = denticon_dump_saturday.strftime('%a %b %-d')  # e.g. "Sat Mar 21"
            reconciliation_note = (
                f"Snapshot taken {snap_sunday_str} at {snap_sunday_time} from Azure SQL dump "
                f"(Denticon data as of {denticon_dump_str} 1am CST). "
                f"This matches Tanya's Income Allocation Report for this period."
            )

            period = fresh.copy()
            period['state']       = 'pending'
            period['locked']      = False
            period['snapshotDate'] = snap['snapshotAt']
            period['reconciliationNote'] = reconciliation_note
            period['snapshot']    = {
                'doctors':    fresh['doctors'],
                'offices':    fresh['offices'],
                'termed':     fresh['termed'],
                'snapshotAt': snap['snapshotAt'],
                'reconciliationNote': reconciliation_note,
            }
            # Remove top-level doctors/offices/termed — they now live in snapshot
            # The HTML reads from snapshot when state = pending
            for key_rm in ('doctors', 'offices', 'termed', 'live'):
                period.pop(key_rm, None)
            periods_out[label] = period
            print(f"    → Pending. Snapshot frozen at {snap['snapshotAt']} — NO further updates.")
            print(f"    Reconciliation note: {reconciliation_note}")
            print(f"    Tanya/Dr. Lugo will Lock when payroll is actually paid.")
            continue

        # ── LIVE: query and refresh ────────────────────────────────────────
        print(f"    State: LIVE — querying fresh data")
        transactions = query_income_allocation(conn, start, end)

        if not transactions:
            print(f"    No data yet for {label} — skipping")
            continue

        if audit_xray:
            audit_xray_codes(transactions)

        period = process_transactions(transactions, start, end, label, pay_date)
        period['state']  = 'live'
        period['locked'] = False
        periods_out[label] = period
        print(f"    LIVE — {len(period['doctors'])} doctors, ${sum(d['coll'] for d in period['doctors']):,.0f} collections")

    conn.close()

    if not periods_out:
        print("ERROR: No period data generated")
        sys.exit(1)

    write_payroll_json(periods_out, output_path)


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def write_payroll_json(periods_dict, output_path):
    """Write payroll.json. Periods sorted chronologically by period_start."""
    # Sort periods chronologically
    def sort_key(item):
        pdata = item[1]
        ps = pdata.get('period_start', '')
        return ps if ps else item[0]

    sorted_periods = dict(sorted(periods_dict.items(), key=sort_key))

    payload = {
        'periods':      sorted_periods,
        'last_updated': datetime.utcnow().isoformat() + 'Z',
        'generated_by': 'payroll_refresh.py',
        'source':       'rpt.vw_income_allocation (Azure SQL) — matches Tanya\'s Income Allocation Detail report',
    }

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(payload, f, indent=2)

    print(f"\n─── Summary ──────────────────────────────────────────────────")
    for key, period in sorted_periods.items():
        state     = period.get('state', period.get('status', '?'))
        hc        = ' *' if period.get('hardcoded') else ''
        snap_info = ''
        if state in ('pending', 'processing') and 'snapshot' in period:
            snap_at = period.get('snapshotDate', period['snapshot'].get('snapshotAt', ''))[:10]
            snap_info = f" (snapped {snap_at})"

        # Use snapshot data for pending/locked periods
        if state in ('pending', 'processing', 'locked') and 'snapshot' in period:
            doctors = period['snapshot'].get('providers') or period['snapshot'].get('doctors', [])
        else:
            doctors = period.get('doctors', [])

        total_coll  = sum(d.get('coll', 0) for d in doctors)
        total_pay   = sum(d.get('payNo', 0) for d in doctors)
        total_saved = sum(d.get('payWith', 0) - d.get('payNo', 0) for d in doctors)

        print(f"  {key:10s} [{state.upper():10s}]{hc}{snap_info}")
        print(f"             coll=${total_coll:>10,.2f}  pay=${total_pay:>9,.2f}  saved=${total_saved:>8,.2f}")

    print(f"\n  Written: {output_path}")
    print(f"  Timestamp: {payload['last_updated']}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='PSPD Payroll Pipeline — generates data/payroll.json with 3-state periods'
    )
    parser.add_argument('--from-excel', '-e',
        help='Path to Denticon Income Allocation Report Excel export')
    parser.add_argument('--output', '-o', default='data/payroll.json',
        help='Output path (default: data/payroll.json)')
    parser.add_argument('--period',
        help='Process a specific period label (e.g. "3.27.26")')
    parser.add_argument('--all-periods', action='store_true',
        help='Process all historical periods (non-locked)')
    parser.add_argument('--recent', action='store_true',
        help='Process last 5 periods (default for Azure task)')
    parser.add_argument('--list-periods', action='store_true',
        help='List available pay periods from database and exit')
    parser.add_argument('--audit-xray', action='store_true',
        help='Run x-ray code audit on a period (use with --period or --from-excel)')

    args = parser.parse_args()

    load_doctor_config()

    # Resolve output path relative to repo root
    output_path = args.output
    if not os.path.isabs(output_path):
        for candidate in ['.', '..', '../..']:
            if os.path.exists(os.path.join(candidate, 'payroll.html')):
                output_path = os.path.join(candidate, output_path)
                break

    conn_str = os.environ.get('AZURE_SQL_CONN_STR', '')

    # ── List periods ────────────────────────────────────────────────────────
    if args.list_periods:
        if not conn_str:
            print("ERROR: Set AZURE_SQL_CONN_STR")
            sys.exit(1)
        conn = get_connection(conn_str)
        periods = get_pay_periods(conn)
        conn.close()
        today = date.today()
        print(f"\nPay Periods ({len(periods)} total):")
        for p in periods[:30]:
            cur = ' ← CURRENT' if p['start'] <= today <= p['end'] else ''
            print(f"  {p['label']:10s}  {p['start']} to {p['end']}  (pay: {p['pay_date']}){cur}")
        if len(periods) > 30:
            print(f"  ...and {len(periods)-30} more")
        return

    # ── Excel mode ──────────────────────────────────────────────────────────
    if args.from_excel:
        filepath = args.from_excel
        if not os.path.exists(filepath):
            print(f"ERROR: File not found: {filepath}")
            sys.exit(1)

        print(f"Reading Excel: {filepath}")
        transactions = parse_excel(filepath)

        if not transactions:
            print("ERROR: No transactions found in Excel file")
            sys.exit(1)

        period_info = find_period_for_transactions(transactions, args.period)
        if period_info is None:
            print("ERROR: Could not detect pay period. Use --period to specify.")
            sys.exit(1)

        label, start, end, pay_date = period_info
        print(f"  Period: {label} ({start} to {end})")

        # Load existing JSON to check lock state
        existing = load_existing_json(output_path)
        if label in existing and is_period_locked(existing[label]):
            print(f"ERROR: Period {label} is LOCKED. Cannot overwrite with Excel data.")
            print("  If you need to correct locked data, unlock it first (Admin panel).")
            sys.exit(1)

        if args.audit_xray:
            audit_xray_codes(transactions)

        period = process_transactions(transactions, start, end, label, pay_date)
        raw_state = get_period_state(end, pay_date)
        # Normalize: use 'pending' (not legacy 'processing')
        period['state']  = 'pending' if raw_state == 'processing' else raw_state
        period['locked'] = False

        # Preserve all existing locked periods, then add/update this one
        periods_out = {}
        for lbl, pdata in existing.items():
            if is_period_locked(pdata):
                periods_out[lbl] = pdata
        # Inject hardcoded
        for lbl, hdata in get_hardcoded_periods().items():
            if lbl not in periods_out:
                periods_out[lbl] = hdata
        periods_out[label] = period
        write_payroll_json(periods_out, output_path)
        return

    # ── Azure SQL mode ──────────────────────────────────────────────────────
    if not conn_str:
        print("ERROR: No data source configured.")
        print("  Use --from-excel <file>  for manual Excel import")
        print("  Or set AZURE_SQL_CONN_STR  for Azure SQL (automated)")
        sys.exit(1)

    if args.all_periods:
        target = 'all'
    elif args.recent:
        target = 'recent'
    elif args.period:
        target = args.period
    else:
        target = 'recent'   # Default: last 5 periods (Azure task default)

    run_azure_pipeline(conn_str, target, output_path, audit_xray=args.audit_xray)


if __name__ == '__main__':
    main()
