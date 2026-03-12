#!/usr/bin/env python3
"""
payroll_refresh.py — PSPD Doctor Payroll Data Pipeline
======================================================
Generates data/payroll.json consumed by payroll.html's loadLiveData() function.

TWO DATA PATHS (automatic failover):
  1. Azure SQL  — queries rpt.vw_income_allocation (primary)
  2. Excel file — reads Tanya's "Income Allocation Report - Detail" export

Azure SQL Views Used:
  rpt.vw_income_allocation           — transaction-level income allocations (Tanya's report)
  rpt.vw_pay_periods                 — pay period definitions (auto-generated through 2037)
  rpt.vw_doctor_collections_no_xray  — collections excluding x-ray procedures

Note: Accuracy is verified manually at payroll time against Tanya's actual
paid amounts, then locked permanently. It is NOT auto-calculated.

Usage:
  # From Azure SQL (automated via GitHub Actions):
  python scripts/payroll_refresh.py

  # From Azure SQL — all historical periods:
  python scripts/payroll_refresh.py --all-periods

  # From Tanya's Excel export (manual fallback):
  python scripts/payroll_refresh.py --from-excel path/to/Denticon_NewMonthlyIncAllD.xlsx

  # Show available pay periods from database:
  python scripts/payroll_refresh.py --list-periods

Environment Variables (for Azure SQL path):
  AZURE_SQL_CONN_STR — Full ODBC connection string for Denticon Azure SQL

Output:
  data/payroll.json — consumed by payroll.html loadLiveData()
"""

import json
import os
import sys
import argparse
import re
from datetime import datetime, date, timedelta
from pathlib import Path

# ---------------------------------------------------------------------------
# CONFIGURATION — edit these when rates change or doctors join/leave
# ---------------------------------------------------------------------------

DOCTOR_CONFIG = {
    # Denticon provider name → display name, pay rate, pay basis, owner flag
    # Provider names must match "Last, First" format from PGID4951_PROVIDER
    # pay_basis: 'collections' (all associates), 'salary' (owner)
    # Schrack is 1099 contract but paid on collections same as W-2 associates
    'Slaven, Chad':       {'display': 'Dr. Slaven',  'pct': 0.36, 'owner': False, 'pay_basis': 'collections'},
    'Menon, Leena':       {'display': 'Dr. Menon',   'pct': 0.35, 'owner': False, 'pay_basis': 'collections'},
    'Choong, Carissa':    {'display': 'Dr. Choong',  'pct': 0.35, 'owner': False, 'pay_basis': 'collections'},
    'Benton, Patricia':   {'display': 'Dr. Benton',  'pct': 0.33, 'owner': False, 'pay_basis': 'collections'},
    'Welter, Erin':       {'display': 'Dr. Welter',  'pct': 0.31, 'owner': False, 'pay_basis': 'collections'},
    'Patel, Dusayant':    {'display': 'Dr. Patel',   'pct': 0.32, 'owner': False, 'pay_basis': 'collections'},
    'Bell, Kendra':       {'display': 'Dr. Bell',    'pct': 0.35, 'owner': False, 'pay_basis': 'collections'},
    'Schrack, Donald':    {'display': 'Dr. Schrack', 'pct': 0.45, 'owner': False, 'pay_basis': 'collections'},
    'Lugo, Christopher':  {'display': 'Dr. Lugo',    'pct': 0.36, 'owner': True,  'pay_basis': 'salary'},
}

# Provider ID → name mapping (from PGID4951_PROVIDER)
# Used by Azure SQL queries that return provider_id instead of name
# These IDs are populated automatically on first run via --build-id-map
PROVIDER_ID_MAP = {}

# Terminated/inactive doctors — tracked but not in active payroll
TERMED_DOCTORS = {
    'Kirk, Kyle':  {'display': 'Dr. Kirk',    'note': 'Terminated'},
    'Ping, Sita':  {'display': 'Dr. Ping',    'note': 'Terminated'},
    'Laws':        {'display': 'Dr. Laws',    'note': 'Terminated'},
}

# Office ID → (display name, abbreviation)
# IDs from PGID4951_OFFICE; update if offices change
OFFICE_ID_MAP = {
    # These will be auto-populated from the database
}

OFFICE_NAME_MAP = {
    # Text matching for Excel path and fallback
    'EVERETT':      ('Everett',      'EV'),
    'LAKE STEVENS': ('Lake Stevens', 'LS'),
    'MARYSVILLE':   ('Marysville',  'MV'),
    'MONROE':       ('Monroe',      'MO'),
    'STANWOOD':     ('Stanwood',    'SW'),
}

# X-ray procedure code prefixes — excluded from payNo calculation
XRAY_PREFIXES = ('D02', 'D03')


# ---------------------------------------------------------------------------
# LOAD DOCTOR CONFIG FROM doctors.json (with hardcoded fallback)
# ---------------------------------------------------------------------------

def load_doctor_config():
    """
    Load doctor configuration from data/doctors.json.
    Falls back to the hardcoded DOCTOR_CONFIG dict if the file is missing.
    Returns (active_doctors_dict, termed_doctors_dict).
    """
    global DOCTOR_CONFIG, TERMED_DOCTORS

    # Try multiple paths (repo root or same dir as script)
    candidates = [
        Path(__file__).resolve().parent.parent / "data" / "doctors.json",
        Path(__file__).resolve().parent / "data" / "doctors.json",
        Path("data/doctors.json"),
    ]

    doctors_json = None
    for p in candidates:
        if p.exists():
            doctors_json = p
            break

    if doctors_json is None:
        print("INFO: data/doctors.json not found — using hardcoded DOCTOR_CONFIG")
        return DOCTOR_CONFIG, TERMED_DOCTORS

    try:
        with open(doctors_json, 'r') as f:
            data = json.load(f)

        # Convert JSON format to DOCTOR_CONFIG format
        active = {}
        for name, doc in data.get("doctors", {}).items():
            active[name] = {
                'display': doc['display'],
                'pct': doc['pct'],
                'owner': doc.get('owner', False),
                'pay_basis': doc.get('pay_basis', 'collections'),
            }

        # Convert terminated
        termed = {}
        for name, doc in data.get("terminated", {}).items():
            termed[name] = {
                'display': doc['display'],
                'note': doc.get('note', 'Terminated'),
            }

        DOCTOR_CONFIG = active
        TERMED_DOCTORS = termed
        print(f"Loaded {len(active)} active + {len(termed)} terminated doctors from {doctors_json}")
        return active, termed

    except Exception as e:
        print(f"WARNING: Failed to load {doctors_json}: {e} — using hardcoded DOCTOR_CONFIG")
        return DOCTOR_CONFIG, TERMED_DOCTORS


# ---------------------------------------------------------------------------
# HARDCODED HISTORICAL DATA — Tanya's verified paid amounts
# ---------------------------------------------------------------------------
# These are the EXACT amounts Tanya paid, extracted from her Master Pay
# Template. They are frozen permanently and marked with hardcoded=True
# so the dashboard shows an asterisk (*) indicating verified historical data.
# The pipeline will NEVER re-query SQL for these periods.

def get_hardcoded_periods():
    """
    Return Tanya's verified historical payroll data.
    Source: 'NEW_MASTER PAY TEMPLATE ASSOCIATE' Excel file, 'No X-Rays' tabs.
    """
    def _fmt(start, end, pay):
        from datetime import date as _d
        s = _d.fromisoformat(start)
        e = _d.fromisoformat(end)
        p = _d.fromisoformat(pay)
        return {
            'dates': f"{s.strftime('%b %-d')} – {e.strftime('%b %-d')}, {e.year}",
            'payDate': p.strftime('%b %-d, %Y'),
            'period_start': start,
            'period_end': end,
            'pay_date_iso': pay,
            'status': 'complete',
            'locked': True,
            'hardcoded': True,
            'locked_date': pay,
            'daysElapsed': (e - s).days + 1,
            'daysTotal': (e - s).days + 1,
        }

    periods = {}

    # ── 1.16.26 (Dec 27, 2025 – Jan 9, 2026) ──
    p = _fmt('2025-12-27', '2026-01-09', '2026-01-16')
    p['label'] = '1.16.26'
    p['doctors'] = [
        {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 40706.06, 'payNo': 14654.18, 'payWith': 17498.80, 'off': {'MV': 15612.50, 'MO': 1022.52, 'SW': 24071.04}},
        {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 36138.64, 'payNo': 11564.36, 'payWith': 13029.92, 'off': {'MV': 8942.95, 'MO': 19031.63, 'LS': 356.06, 'EV': 7808.00}},
        {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 33875.44, 'payNo': 11856.40, 'payWith': 14055.46, 'off': {'MV': 10324.63, 'SW': 23449.65, 'EV': 101.16}},
        {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 30738.98, 'payNo': 10758.64, 'payWith': 12472.02, 'off': {'MV': 4357.82, 'MO': 20110.41, 'EV': 6270.75}},
        {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 27164.96, 'payNo': 8421.14,  'payWith': 10119.69, 'off': {'MV': 13678.80, 'MO': 13269.16, 'SW': 217.00}},
        {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 21251.67, 'payNo': 7438.08,  'payWith': 8879.72,  'off': {'MV': 670.20, 'MO': 19461.27, 'EV': 1120.20}},
        {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 450.70,   'payNo': 144.22,   'payWith': 175.90,   'off': {'MV': 101.20, 'MO': 349.50}},
        {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 0.00,     'payNo': 0.00,     'payWith': 0.00},
    ]
    p['offices'] = [
        {'name': 'Monroe',       'amt': 82545.87},
        {'name': 'Marysville',   'amt': 58154.50},
        {'name': 'Stanwood',     'amt': 47737.69},
        {'name': 'Everett',      'amt': 30446.41},
        {'name': 'Lake Stevens', 'amt': 356.06},
    ]
    p['termed'] = [{'name': 'Dr. Kirk', 'coll': 28914.08, 'note': 'Terminated'}]
    periods['1.16.26'] = p

    # ── 1.30.26 (Jan 10 – Jan 23, 2026) ──
    p = _fmt('2026-01-10', '2026-01-23', '2026-01-30')
    p['label'] = '1.30.26'
    p['doctors'] = [
        {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 57272.45, 'payNo': 20618.08, 'payWith': 23550.87, 'off': {'MV': 22683.98, 'MO': 7543.50, 'SW': 27044.97}},
        {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 52757.04, 'payNo': 18464.96, 'payWith': 21338.47, 'off': {'MV': 10645.82, 'MO': 30920.14, 'LS': 4514.09, 'EV': 6676.99}},
        {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 46867.49, 'payNo': 14997.60, 'payWith': 17744.43, 'off': {'MV': 21326.45, 'MO': 22960.86, 'LS': 463.40, 'EV': 2116.78}},
        {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 42240.48, 'payNo': 14784.17, 'payWith': 17380.04, 'off': {'MV': 20835.35, 'SW': 21071.26, 'EV': 333.87}},
        {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 29843.79, 'payNo': 9251.57,  'payWith': 10988.61, 'off': {'MV': 16268.65, 'MO': 3955.53, 'LS': 2636.10, 'SW': 56.94, 'EV': 6926.57}},
        {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 21932.74, 'payNo': 7676.46,  'payWith': 8943.45,  'off': {'MV': 527.60, 'MO': 16752.57, 'LS': 1850.75, 'EV': 2801.82}},
        {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 1642.76,  'payNo': 739.24,   'payWith': 739.24,   'off': {'MV': 1491.16, 'EV': 151.60}},
        {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 445.10,   'payNo': 142.43,   'payWith': 147.17,   'off': {'MO': 445.10}},
    ]
    p['offices'] = [
        {'name': 'Marysville',   'amt': 100800.74},
        {'name': 'Monroe',       'amt': 84035.16},
        {'name': 'Stanwood',     'amt': 48842.88},
        {'name': 'Everett',      'amt': 26515.45},
        {'name': 'Lake Stevens', 'amt': 10345.60},
    ]
    p['termed'] = [{'name': 'Dr. Kirk', 'coll': 17537.98, 'note': 'Terminated'}]
    periods['1.30.26'] = p

    # ── 2.13.26 (Jan 24 – Feb 6, 2026) — NOTE: Benton rate = 0.33 this period ──
    p = _fmt('2026-01-24', '2026-02-06', '2026-02-13')
    p['label'] = '2.13.26'
    p['doctors'] = [
        {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 70435.18, 'payNo': 25356.66, 'payWith': 29323.52, 'off': {'MV': 34056.54, 'MO': 2468.80, 'SW': 33909.84}},
        {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 54346.07, 'payNo': 16847.28, 'payWith': 18518.56, 'off': {'MV': 9983.65, 'MO': 5563.26, 'LS': 22371.29, 'SW': 937.98, 'EV': 15489.89}},
        {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 52215.52, 'payNo': 18275.43, 'payWith': 21036.47, 'off': {'MV': 23695.69, 'MO': 3615.50, 'SW': 24904.33}},
        {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 43717.91, 'payNo': 15301.27, 'payWith': 17589.10, 'off': {'MV': 268.32, 'MO': 24897.54, 'LS': 12179.74, 'EV': 6372.31}},
        {'name': 'Dr. Benton',  'pct': 0.33, 'coll': 38867.69, 'payNo': 12826.34, 'payWith': 14665.29, 'off': {'MV': 7224.62, 'MO': 24637.62, 'LS': 6652.45, 'EV': 353.00}},
        {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 24586.75, 'payNo': 8605.36,  'payWith': 9617.71,  'off': {'MV': 258.60, 'MO': 12073.41, 'LS': 12041.77, 'EV': 212.97}},
        {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 5443.12,  'payNo': 1741.80,  'payWith': 2141.82,  'off': {'MV': 1376.24, 'MO': 4066.88}},
        {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 2023.41,  'payNo': 910.53,   'payWith': 910.53,   'off': {'MV': 1544.01, 'EV': 479.40}},
    ]
    p['offices'] = [
        {'name': 'Marysville',   'amt': 78477.67},
        {'name': 'Monroe',       'amt': 77629.78},
        {'name': 'Stanwood',     'amt': 59796.15},
        {'name': 'Lake Stevens', 'amt': 53245.25},
        {'name': 'Everett',      'amt': 23858.82},
    ]
    p['termed'] = [{'name': 'Dr. Kirk', 'coll': 1372.02, 'note': 'Terminated'}]
    periods['2.13.26'] = p

    # ── 2.27.26 (Feb 7 – Feb 20, 2026) — NOTE: Benton rate back to 0.32 ──
    p = _fmt('2026-02-07', '2026-02-20', '2026-02-27')
    p['label'] = '2.27.26'
    p['doctors'] = [
        {'name': 'Dr. Slaven',  'pct': 0.36, 'coll': 68512.61, 'payNo': 24664.54, 'payWith': 28572.15, 'off': {'MV': 27086.51, 'MO': 11879.86, 'SW': 29546.24}},
        {'name': 'Dr. Menon',   'pct': 0.35, 'coll': 49682.98, 'payNo': 17389.04, 'payWith': 19276.69, 'off': {'MV': 157.00, 'MO': 20429.10, 'LS': 14883.05, 'EV': 14213.83}},
        {'name': 'Dr. Choong',  'pct': 0.35, 'coll': 47354.85, 'payNo': 16574.20, 'payWith': 18910.47, 'off': {'MV': 19626.25, 'MO': 1889.00, 'SW': 25839.60}},
        {'name': 'Dr. Benton',  'pct': 0.32, 'coll': 43720.53, 'payNo': 13990.57, 'payWith': 15844.07, 'off': {'MV': 12134.16, 'MO': 18453.36, 'LS': 13133.01}},
        {'name': 'Dr. Welter',  'pct': 0.31, 'coll': 38646.25, 'payNo': 11980.34, 'payWith': 13737.56, 'off': {'MV': 5937.86, 'MO': 168.96, 'LS': 20983.32, 'SW': 4166.44, 'EV': 7389.67}},
        {'name': 'Dr. Patel',   'pct': 0.32, 'coll': 31620.97, 'payNo': 10118.71, 'payWith': 11718.98, 'off': {'MV': 7235.48, 'MO': 24385.49}},
        {'name': 'Dr. Bell',    'pct': 0.35, 'coll': 25535.53, 'payNo': 8937.44,  'payWith': 10114.36, 'off': {'MV': 41.80, 'MO': 10481.83, 'LS': 15011.90}},
        {'name': 'Dr. Schrack', 'pct': 0.45, 'coll': 7528.70,  'payNo': 3387.91,  'payWith': 3387.91,  'off': {'MV': 4113.00, 'EV': 3415.70}},
    ]
    p['offices'] = [
        {'name': 'Monroe',       'amt': 87687.60},
        {'name': 'Marysville',   'amt': 76332.06},
        {'name': 'Lake Stevens', 'amt': 64011.28},
        {'name': 'Stanwood',     'amt': 59552.28},
        {'name': 'Everett',      'amt': 25019.20},
    ]
    p['termed'] = [{'name': 'Dr. Kirk', 'coll': 0.00, 'note': 'Terminated'}]
    periods['2.27.26'] = p

    return periods


# ---------------------------------------------------------------------------
# AZURE SQL — queries the real rpt.* views
# ---------------------------------------------------------------------------

def get_connection(conn_str):
    """Create Azure SQL connection."""
    try:
        import pyodbc
    except ImportError:
        print("ERROR: pyodbc required. Install with: pip install pyodbc")
        sys.exit(1)

    # Show available drivers for debugging
    drivers = pyodbc.drivers()
    print(f"  Available ODBC drivers: {drivers}")

    # Validate connection string has required components
    cs_lower = conn_str.lower()
    if 'driver' not in cs_lower:
        print("WARNING: Connection string missing 'Driver=' parameter")
        print("  Expected format: Driver={ODBC Driver 18 for SQL Server};Server=tcp:YOUR-SERVER.database.windows.net,1433;Database=YOUR-DB;Uid=YOUR-USER;Pwd=YOUR-PASS;Encrypt=yes;TrustServerCertificate=no;Connection Timeout=30;")
    if 'your-' in cs_lower or 'your_' in cs_lower:
        print("ERROR: Connection string contains placeholder values (YOUR-SERVER, YOUR-DATABASE, etc.)")
        print("  Replace these with your actual Azure SQL credentials")
        sys.exit(1)

    try:
        conn = pyodbc.connect(conn_str, timeout=30)
        print("  Connected to Azure SQL successfully")
        return conn
    except pyodbc.Error as e:
        print(f"ERROR: Failed to connect to Azure SQL")
        print(f"  Error: {e}")
        print(f"  Connection string length: {len(conn_str)} chars")
        # Show masked server name for debugging
        import re
        server_match = re.search(r'Server\s*=\s*tcp:([^,;]+)', conn_str, re.IGNORECASE)
        if server_match:
            print(f"  Server: {server_match.group(1)}")
        db_match = re.search(r'Database\s*=\s*([^;]+)', conn_str, re.IGNORECASE)
        if db_match:
            print(f"  Database: {db_match.group(1)}")
        sys.exit(1)


def load_id_maps(conn):
    """Load provider and office ID→name mappings from Denticon tables."""
    cursor = conn.cursor()

    # Provider ID → "Last, First"
    cursor.execute("""
        SELECT PROVIDERID, LNAME, FNAME
        FROM PGID4951_PROVIDER
        WHERE LNAME IS NOT NULL
    """)
    for r in cursor.fetchall():
        name = f"{r.LNAME.strip()}, {r.FNAME.strip()}"
        PROVIDER_ID_MAP[r.PROVIDERID] = name

    # Office ID → name
    cursor.execute("""
        SELECT OID, OFFICENAME
        FROM PGID4951_OFFICE
    """)
    for r in cursor.fetchall():
        office_name = r.OFFICENAME.strip() if r.OFFICENAME else f"Office {r.OID}"
        OFFICE_ID_MAP[r.OID] = office_name

    print(f"  Loaded {len(PROVIDER_ID_MAP)} providers, {len(OFFICE_ID_MAP)} offices")


def get_pay_periods(conn):
    """Fetch pay period definitions from rpt.vw_pay_periods."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT pay_period_num, period_start, period_end
        FROM rpt.vw_pay_periods
        ORDER BY period_start
    """)
    periods = []
    for r in cursor.fetchall():
        start = r.period_start if isinstance(r.period_start, date) else date.fromisoformat(str(r.period_start)[:10])
        end = r.period_end if isinstance(r.period_end, date) else date.fromisoformat(str(r.period_end)[:10])
        # Generate label: pay date is 7 days after period end (biweekly)
        pay_date = end + timedelta(days=7)
        label = f"{pay_date.month}.{pay_date.day}.{str(pay_date.year)[2:]}"
        periods.append({
            'num': r.pay_period_num,
            'start': start,
            'end': end,
            'pay_date': pay_date,
            'label': label,
        })
    return periods


def find_current_period(periods):
    """Find the current (or most recent) pay period."""
    today = date.today()
    # First: find a period containing today
    for p in periods:
        if p['start'] <= today <= p['end']:
            return p
    # Fallback: find the most recent completed period
    past = [p for p in periods if p['end'] < today]
    if past:
        return past[-1]
    # Last resort: first future period
    return periods[0] if periods else None


def query_income_allocation_azure(conn, period_start, period_end):
    """
    Query rpt.vw_income_allocation for transaction-level income data.
    This is the same data as Tanya's "Income Allocation Report - Detail".

    Returns list of transaction dicts matching the Excel parser output format.
    """
    cursor = conn.cursor()

    cursor.execute("""
        SELECT
            alloc_provider_id,
            OID AS office_id,
            ALLOCDATE,
            alloc_amount,
            proc_ada_code,
            period_label
        FROM rpt.vw_income_allocation
        WHERE period_start = ? AND period_end = ?
    """, (str(period_start), str(period_end)))

    transactions = []
    for r in cursor.fetchall():
        provider_name = PROVIDER_ID_MAP.get(r.alloc_provider_id, f"Provider {r.alloc_provider_id}")
        office_name = OFFICE_ID_MAP.get(r.office_id, f"Office {r.office_id}")

        # Normalize office name for matching
        office_key = office_name.upper().replace('PSPD - ', '').replace('PSPD-', '').strip()

        transactions.append({
            'office': office_key,
            'provider': provider_name,
            'alloc_date': r.ALLOCDATE,
            'proc_code': str(r.proc_ada_code).strip() if r.proc_ada_code else '',
            'income': float(r.alloc_amount) if r.alloc_amount else 0.0,
        })

    print(f"    vw_income_allocation: {len(transactions)} rows")
    return transactions


def load_existing_json(output_path):
    """Load existing payroll.json to preserve locked periods."""
    try:
        with open(output_path, 'r') as f:
            data = json.load(f)
        return data.get('periods', {})
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def is_period_locked(period_data):
    """Check if a period has been locked (finalized after payroll)."""
    return period_data.get('locked', False)


def should_lock_period(period_end_obj):
    """
    Data locks on the Sunday after the period ends (period_end + 2 days).
    Timeline:
      - Friday (period_end): Last day of pay period
      - Saturday (period_end + 1): Denticon → Azure SQL data dump
      - Sunday AM (period_end + 2): Pipeline locks data permanently
      - Sun PM / Monday: Tanya reviews locked numbers (Processing)
      - Friday (pay_date, period_end + 7): Payroll paid (Complete)
    """
    today = date.today()
    lock_date = period_end_obj + timedelta(days=2)
    return today >= lock_date


def get_period_status(period_end_obj, pay_date_obj):
    """
    Three-phase period lifecycle:
      - 'live'       : today <= period_end (period active, data recalculates)
      - 'processing' : period_end < today < pay_date (data locked, Tanya prepping)
      - 'complete'   : today >= pay_date (payroll paid, done)
    """
    today = date.today()
    if today <= period_end_obj:
        return 'live'
    elif today < pay_date_obj:
        return 'processing'
    else:
        return 'complete'


def run_azure_pipeline(conn_str, target_periods, output_path):
    """
    Full Azure SQL pipeline with PERIOD LOCKING:
    1. Load existing payroll.json to find already-locked periods
    2. Load ID maps (provider/office names)
    3. Get pay period definitions from DB
    4. For each target period:
       - If LOCKED in existing JSON → keep as-is, skip query
       - If past period_end+2 (Sunday) and not locked → query, calculate, LOCK
       - If LIVE (current period) → query and recalculate fresh
    5. Assign three-phase status: live → processing → complete
    6. Merge and write payroll.json
    """
    # Step 0: Load hardcoded historical periods (Tanya's verified data)
    hardcoded = get_hardcoded_periods()
    print(f"  Hardcoded historical periods: {', '.join(sorted(hardcoded.keys()))}")

    # Step 1: Load existing data to preserve locked periods
    existing_periods = load_existing_json(output_path)
    locked_count = sum(1 for p in existing_periods.values() if is_period_locked(p))
    print(f"  Existing payroll.json: {len(existing_periods)} periods ({locked_count} locked)")

    conn = get_connection(conn_str)

    print("  Loading provider/office ID maps...")
    load_id_maps(conn)

    print("  Fetching pay periods from rpt.vw_pay_periods...")
    all_periods = get_pay_periods(conn)
    print(f"    Found {len(all_periods)} pay periods ({all_periods[0]['label']} to {all_periods[-1]['label']})")

    # Determine which periods to process
    if target_periods == 'current':
        current = find_current_period(all_periods)
        if not current:
            print("ERROR: Could not determine current pay period")
            sys.exit(1)
        periods_to_process = [current]
    elif target_periods == 'all':
        today = date.today()
        periods_to_process = [p for p in all_periods if p['start'] <= today]
    elif target_periods == 'recent':
        # Last 5 periods (for dashboard history)
        today = date.today()
        past = [p for p in all_periods if p['start'] <= today]
        periods_to_process = past[-5:] if len(past) >= 5 else past
    else:
        # Specific period label
        match = [p for p in all_periods if p['label'] == target_periods]
        if not match:
            print(f"ERROR: Period '{target_periods}' not found in database")
            print(f"  Available: {', '.join(p['label'] for p in all_periods[:20])}...")
            sys.exit(1)
        periods_to_process = match

    # Start with ALL existing locked periods preserved
    periods_dict = {}
    for label, pdata in existing_periods.items():
        if is_period_locked(pdata):
            periods_dict[label] = pdata

    # Inject hardcoded periods (always take priority — Tanya's verified data)
    for label, hdata in hardcoded.items():
        if label not in periods_dict:
            periods_dict[label] = hdata
            print(f"  Injected hardcoded period: {label} *")
        elif not periods_dict[label].get('hardcoded'):
            # Replace SQL-derived data with Tanya's verified data
            periods_dict[label] = hdata
            print(f"  Replaced with hardcoded period: {label} *")

    # Process each target period
    for period_info in periods_to_process:
        label = period_info['label']
        start = period_info['start']
        end = period_info['end']
        pay_date = period_info['pay_date']

        # Check if already locked (including hardcoded) — skip entirely
        if label in periods_dict and is_period_locked(periods_dict[label]):
            hc = " (hardcoded *)" if periods_dict[label].get('hardcoded') else ""
            print(f"\n  Period {label}: LOCKED{hc} — skipping")
            continue

        print(f"\n  Processing period {label} ({start} to {end})...")

        # Query fresh data from SQL
        transactions = query_income_allocation_azure(conn, start, end)

        if not transactions:
            print(f"    No data for period {label} — skipping")
            continue

        # Process transactions into dashboard format
        period = process_transactions(transactions, start, end, label, pay_date)

        # Determine three-phase status
        status = get_period_status(end, pay_date)
        period['status'] = status

        # Lock data on the Sunday after period ends (period_end + 2)
        if should_lock_period(end):
            period['locked'] = True
            period['locked_date'] = date.today().isoformat()
            print(f"    LOCKED — data frozen (period ended {end}, dump Sat, locked Sun)")
            print(f"    Status: {status.upper()}")
        else:
            period['locked'] = False
            print(f"    LIVE — data will recalculate on next run")

        periods_dict[label] = period

    conn.close()

    if not periods_dict:
        print("ERROR: No data retrieved for any period")
        sys.exit(1)

    write_payroll_json(periods_dict, output_path)


# ---------------------------------------------------------------------------
# EXCEL PARSING — reads Tanya's "Income Allocation Report - Detail"
# ---------------------------------------------------------------------------

def parse_excel(filepath):
    """
    Parse Denticon 'Income Allocation Report - Detail' Excel export.
    Returns list of transaction dicts with: office, provider, alloc_date,
    proc_code, income.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    transactions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        office = None
        current_provider = None

        for row in ws.iter_rows(min_row=1, values_only=False):
            a_val = row[0].value if len(row) > 0 else None
            b_val = row[1].value if len(row) > 1 else None
            c_val = row[2].value if len(row) > 2 else None

            # Detect office: "Office: PSPD - EVERETT"
            if a_val and isinstance(a_val, str) and 'Office:' in a_val:
                match = re.search(r'PSPD\s*-\s*(.+)', a_val)
                if match:
                    office = match.group(1).strip().upper()
                continue

            # Detect provider: "Provider :- Bell, Kendra  DDS  : BELL"
            if b_val and isinstance(b_val, str) and 'Provider :-' in b_val:
                parts = b_val.split(':-')[1].strip()
                provider_match = re.match(r'([^,]+,\s*\S+)', parts)
                if provider_match:
                    current_provider = provider_match.group(1).strip()
                continue

            # Transaction rows have datetime in column C (Alloc Date)
            if c_val and isinstance(c_val, datetime) and current_provider and office:
                income = row[14].value if len(row) > 14 else None  # Column O
                proc_code = row[12].value if len(row) > 12 else None  # Column M

                if income is not None:
                    transactions.append({
                        'office': office,
                        'provider': current_provider,
                        'alloc_date': c_val,
                        'proc_code': str(proc_code).strip() if proc_code else '',
                        'income': float(income),
                    })

    wb.close()
    return transactions


# ---------------------------------------------------------------------------
# PROCESSING — transform transactions into payroll.json format
# ---------------------------------------------------------------------------

# Hardcoded pay periods as fallback for Excel mode (when DB not available)
FALLBACK_PAY_PERIODS = [
    ('1.16.26',  '2025-12-27', '2026-01-09', '2026-01-16'),
    ('1.30.26',  '2026-01-10', '2026-01-23', '2026-01-30'),
    ('2.13.26',  '2026-01-24', '2026-02-06', '2026-02-13'),
    ('2.27.26',  '2026-02-07', '2026-02-20', '2026-02-27'),
    ('3.13.26',  '2026-02-21', '2026-03-06', '2026-03-13'),
    ('3.27.26',  '2026-03-07', '2026-03-20', '2026-03-27'),
    ('4.10.26',  '2026-03-21', '2026-04-03', '2026-04-10'),
    ('4.24.26',  '2026-04-04', '2026-04-17', '2026-04-24'),
]


def find_pay_period_fallback(target_date=None):
    """Find pay period from hardcoded list (for Excel mode)."""
    if target_date is None:
        target_date = date.today()
    for label, start, end, pay in FALLBACK_PAY_PERIODS:
        s = date.fromisoformat(start)
        e = date.fromisoformat(end)
        if s <= target_date <= e:
            return label, s, e, date.fromisoformat(pay)
    label, start, end, pay = FALLBACK_PAY_PERIODS[-1]
    return label, date.fromisoformat(start), date.fromisoformat(end), date.fromisoformat(pay)


def determine_period_for_transactions(transactions):
    """Determine which pay period transactions belong to based on alloc dates."""
    if not transactions:
        return None
    dates = [t['alloc_date'] for t in transactions if isinstance(t['alloc_date'], datetime)]
    if not dates:
        return None
    min_date = min(dates).date()
    max_date = max(dates).date()
    for label, start, end, pay in FALLBACK_PAY_PERIODS:
        s = date.fromisoformat(start)
        e = date.fromisoformat(end)
        if min_date <= e and max_date >= s:
            return label, s, e, date.fromisoformat(pay)
    return None


def process_transactions(transactions, period_start, period_end, label, pay_date):
    """
    Transform raw transactions into the payroll.json format that
    payroll.html's loadLiveData() expects.
    """
    today = date.today()
    days_total = (period_end - period_start).days + 1
    days_elapsed = max(1, min(days_total, (today - period_start).days + 1))

    # Aggregate by provider and office
    provider_data = {}
    office_totals = {}

    # IMPORTANT: Denticon Income values are SIGNED.
    # Payments received = negative, adjustments = positive.
    # We sum raw values (preserving sign) then take abs() of the NET total.
    # This matches Tanya's report which shows abs(sum), NOT sum(abs).

    raw_provider = {}
    raw_office = {}

    for t in transactions:
        prov = t['provider']
        office = t['office']
        income = t['income']
        proc_code = t['proc_code']
        is_xray = any(proc_code.startswith(p) for p in XRAY_PREFIXES)

        if prov not in raw_provider:
            raw_provider[prov] = {'total': 0.0, 'xray': 0.0, 'offices': {}}
        raw_provider[prov]['total'] += income
        if is_xray:
            raw_provider[prov]['xray'] += income
        if office not in raw_provider[prov]['offices']:
            raw_provider[prov]['offices'][office] = 0.0
        raw_provider[prov]['offices'][office] += income

        if office not in raw_office:
            raw_office[office] = 0.0
        raw_office[office] += income

    # Convert to absolute values (abs of net sum per provider/office)
    for prov, data in raw_provider.items():
        provider_data[prov] = {
            'coll': abs(data['total']),
            'xray': abs(data['xray']),
            'offices': {k: abs(v) for k, v in data['offices'].items()},
        }
    for office, total in raw_office.items():
        office_totals[office] = abs(total)

    # Build doctor array (active doctors)
    doctors = []
    for denticon_name, config in DOCTOR_CONFIG.items():
        pdata = provider_data.get(denticon_name, {'coll': 0.0, 'xray': 0.0, 'offices': {}})
        coll = round(pdata['coll'], 2)
        xray = round(pdata['xray'], 2)
        rate = config['pct']

        pay_with = round(coll * rate, 2)
        pay_no = 0.0 if config['owner'] else round((coll - xray) * rate, 2)

        doc_entry = {
            'name': config['display'],
            'pct': rate,
            'coll': coll,
            'payNo': pay_no,
            'payWith': pay_with,
        }

        # Per-office breakdown
        off = {}
        for raw_office_name, amount in pdata['offices'].items():
            for office_key, (_, abbr) in OFFICE_NAME_MAP.items():
                if office_key in raw_office_name.upper():
                    off[abbr] = round(amount, 2)
                    break
        if off:
            doc_entry['off'] = off

        doctors.append(doc_entry)

    doctors.sort(key=lambda d: d['coll'], reverse=True)

    # Build office array
    offices = []
    for raw_office_name, total in sorted(office_totals.items(), key=lambda x: -x[1]):
        display_name = raw_office_name
        for office_key, (disp, _) in OFFICE_NAME_MAP.items():
            if office_key in raw_office_name.upper():
                display_name = disp
                break
        offices.append({'name': display_name, 'amt': round(total, 2)})

    # Build termed array
    termed = []
    for denticon_name, config in TERMED_DOCTORS.items():
        pdata = provider_data.get(denticon_name, {'coll': 0.0})
        termed.append({
            'name': config['display'],
            'coll': round(pdata['coll'], 2),
            'note': config['note'],
        })

    # Format dates
    def fmt_date(d):
        return d.strftime('%b %-d, %Y')
    def fmt_date_short(d):
        return d.strftime('%b %-d')

    period = {
        'label': label,
        'dates': f"{fmt_date_short(period_start)} – {fmt_date_short(period_end)}, {period_end.year}",
        'payDate': fmt_date(pay_date),
        'period_start': str(period_start),   # ISO date for client-side auto-advance
        'period_end': str(period_end),        # ISO date for client-side auto-advance
        'pay_date_iso': str(pay_date),        # ISO date for client-side closed detection
        'status': get_period_status(period_end, pay_date),
        'daysElapsed': days_elapsed,
        'daysTotal': days_total,
        'doctors': doctors,
        'offices': offices,
        'termed': termed,
    }

    return period


# ---------------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------------

def write_payroll_json(periods_dict, output_path):
    """Write the payroll.json file that payroll.html consumes."""
    payload = {
        'periods': periods_dict,
        'last_updated': datetime.utcnow().isoformat() + 'Z',
        'generated_by': 'payroll_refresh.py',
        'source': 'rpt.vw_income_allocation (Azure SQL)',
    }

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(payload, f, indent=2)

    # Print summary
    for key, period in sorted(periods_dict.items()):
        total_coll = sum(d['coll'] for d in period['doctors'])
        total_pay = sum(d['payNo'] for d in period['doctors'])
        total_saved = sum(d['payWith'] - d['payNo'] for d in period['doctors'])
        lock_icon = "LOCKED" if period.get('locked') else "LIVE"
        hc_marker = " *" if period.get('hardcoded') else ""
        lock_date = f" (locked {period.get('locked_date', '')})" if period.get('locked') else ""
        print(f"\n  Period {key} [{lock_icon}]{hc_marker}{lock_date} ({period['status']}):")
        print(f"    Collections: ${total_coll:,.2f}")
        print(f"    Pay (no xray): ${total_pay:,.2f}")
        print(f"    X-Ray Savings: ${total_saved:,.2f}")
        print(f"    Offices: {len(period['offices'])}")
        print(f"    Active Doctors: {len(period['doctors'])}")
        for d in period['doctors']:
            saved = d['payWith'] - d['payNo']
            print(f"      {d['name']:14s}  coll=${d['coll']:>10,.2f}  payNo=${d['payNo']:>9,.2f}  saved=${saved:>8,.2f}")

    print(f"\n  Written to: {output_path}")
    print(f"  Timestamp: {payload['last_updated']}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='PSPD Payroll Data Pipeline — generates data/payroll.json'
    )
    parser.add_argument('--from-excel', '-e',
        help='Path to Denticon Income Allocation Report Excel file')
    parser.add_argument('--output', '-o', default='data/payroll.json',
        help='Output path (default: data/payroll.json)')
    parser.add_argument('--period',
        help='Specific pay period label (e.g. "3.13.26"). Default: current.')
    parser.add_argument('--all-periods', action='store_true',
        help='Generate all historical periods')
    parser.add_argument('--recent', action='store_true',
        help='Generate last 5 periods (default for automated runs)')
    parser.add_argument('--list-periods', action='store_true',
        help='List available pay periods from database')
    parser.add_argument('--discover-views', action='store_true',
        help='List all views/tables in the database (find production views for Schrack)')

    args = parser.parse_args()

    # Load doctor config from doctors.json (falls back to hardcoded)
    load_doctor_config()

    # Resolve output path relative to repo root
    output_path = args.output
    if not os.path.isabs(output_path):
        for candidate in ['.', '..', '../..']:
            if os.path.exists(os.path.join(candidate, 'payroll.html')):
                output_path = os.path.join(candidate, output_path)
                break

    conn_str = os.environ.get('AZURE_SQL_CONN_STR', '')

    # --- List periods ---
    if args.list_periods:
        if not conn_str:
            print("ERROR: Set AZURE_SQL_CONN_STR")
            sys.exit(1)
        conn = get_connection(conn_str)
        periods = get_pay_periods(conn)
        conn.close()
        today = date.today()
        print(f"\nPay Periods ({len(periods)} total):")
        for p in periods[:30]:
            marker = ' ← CURRENT' if p['start'] <= today <= p['end'] else ''
            marker = marker or (' ← NEXT' if p['start'] > today and not any(
                pp['start'] <= today <= pp['end'] for pp in periods) else '')
            print(f"  {p['label']:10s}  {p['start']} to {p['end']}  (pay: {p['pay_date']}){marker}")
        if len(periods) > 30:
            print(f"  ... and {len(periods) - 30} more")
        return

    # --- Discover views mode ---
    if args.discover_views:
        if not conn_str:
            print("ERROR: Set AZURE_SQL_CONN_STR")
            sys.exit(1)
        conn = get_connection(conn_str)
        cursor = conn.cursor()

        print("\n=== ALL VIEWS (rpt.* schema) ===")
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME
            FROM INFORMATION_SCHEMA.VIEWS
            WHERE TABLE_SCHEMA = 'rpt'
            ORDER BY TABLE_NAME
        """)
        for row in cursor.fetchall():
            print(f"  {row.TABLE_SCHEMA}.{row.TABLE_NAME}")

        print("\n=== ALL VIEWS (all schemas) ===")
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME
            FROM INFORMATION_SCHEMA.VIEWS
            ORDER BY TABLE_SCHEMA, TABLE_NAME
        """)
        for row in cursor.fetchall():
            print(f"  {row.TABLE_SCHEMA}.{row.TABLE_NAME}")

        # Search specifically for production-related views
        print("\n=== VIEWS/TABLES containing 'produc' ===")
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME, TABLE_TYPE
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME LIKE '%produc%'
            ORDER BY TABLE_SCHEMA, TABLE_NAME
        """)
        results = cursor.fetchall()
        if results:
            for row in results:
                print(f"  {row.TABLE_SCHEMA}.{row.TABLE_NAME} ({row.TABLE_TYPE})")
        else:
            print("  (none found)")

        # Also check for anything with 'sched' or 'appt' that might track production
        print("\n=== VIEWS/TABLES containing 'sched', 'appt', 'proc', 'charge' ===")
        cursor.execute("""
            SELECT TABLE_SCHEMA, TABLE_NAME, TABLE_TYPE
            FROM INFORMATION_SCHEMA.TABLES
            WHERE TABLE_NAME LIKE '%sched%'
               OR TABLE_NAME LIKE '%appt%'
               OR TABLE_NAME LIKE '%proc%'
               OR TABLE_NAME LIKE '%charge%'
            ORDER BY TABLE_SCHEMA, TABLE_NAME
        """)
        results = cursor.fetchall()
        if results:
            for row in results:
                print(f"  {row.TABLE_SCHEMA}.{row.TABLE_NAME} ({row.TABLE_TYPE})")
        else:
            print("  (none found)")

        conn.close()
        return

    # --- Excel mode ---
    if args.from_excel:
        filepath = args.from_excel
        if not os.path.exists(filepath):
            print(f"ERROR: File not found: {filepath}")
            sys.exit(1)

        print(f"Reading Excel: {filepath}")
        transactions = parse_excel(filepath)
        print(f"  Parsed {len(transactions)} transaction rows")

        if not transactions:
            print("ERROR: No transactions found")
            sys.exit(1)

        period_info = determine_period_for_transactions(transactions)
        if period_info is None and args.period:
            for label, start, end, pay in FALLBACK_PAY_PERIODS:
                if label == args.period:
                    period_info = (label, date.fromisoformat(start),
                                   date.fromisoformat(end), date.fromisoformat(pay))
                    break
        if period_info is None:
            print("ERROR: Could not detect pay period. Use --period to specify.")
            sys.exit(1)

        label, start, end, pay_date = period_info
        print(f"  Pay period: {label} ({start} to {end})")
        period = process_transactions(transactions, start, end, label, pay_date)
        write_payroll_json({label: period}, output_path)
        return

    # --- Azure SQL mode ---
    if not conn_str:
        print("ERROR: No data source specified.")
        print("  Use --from-excel <file> for Excel import")
        print("  Or set AZURE_SQL_CONN_STR for Azure SQL")
        sys.exit(1)

    print("Connecting to Azure SQL...")

    if args.all_periods:
        target = 'all'
    elif args.recent:
        target = 'recent'
    elif args.period:
        target = args.period
    else:
        target = 'recent'  # Default: last 5 periods

    run_azure_pipeline(conn_str, target, output_path)


if __name__ == '__main__':
    main()
